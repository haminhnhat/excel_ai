from __future__ import annotations

import shutil
import tempfile
import zipfile
from xml.etree import ElementTree as ET
from pathlib import Path
from typing import Any
import warnings

import yaml
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries

from .utils import normalize_text, safe_filename

FRIENDLY_NAMES = {
    "loan_interest_rate": "Lãi vay ngân hàng",
    "vat_rate": "Thuế VAT",
    "cit_rate": "Thuế TNDN / CIT",
    "selling_price_change": "Điều chỉnh giá bán",
    "investment_cost_change": "Điều chỉnh tổng mức đầu tư / TMĐT",
    "selling_cost_rate": "Chi phí bán hàng / môi giới",
    "marketing_rate": "Chi phí marketing",
    "admin_cost_rate": "Chi phí quản lý",
    "sales_volume_change": "Điều chỉnh sản lượng bán",
    "debt_financing_ratio": "Tỷ lệ vay / debt ratio",
    "project_npv": "NPV dự án",
    "project_irr": "IRR dự án",
    "equity_npv": "NPV chủ đầu tư",
    "equity_irr": "IRR chủ đầu tư",
    "roi": "ROI",
    "profit_after_tax": "Lợi nhuận sau thuế",
    "total_investment": "Tổng mức đầu tư",
    "bank_loan": "Vay ngân hàng",
    "total_revenue": "Tổng doanh thu",
    "revenue": "Doanh thu",
}

INPUT_KEYWORDS: dict[str, list[str]] = {
    "loan_interest_rate": ["lãi vay", "lai vay", "lãi suất", "lai suat", "lãi ngân hàng", "chi phí vốn vay", "interest rate", "loan interest"],
    "vat_rate": ["vat", "thuế vat", "thue vat"],
    "cit_rate": ["tndn", "thuế tndn", "thue tndn", "cit", "corporate income tax"],
    "selling_price_change": ["giá bán", "gia ban", "đơn giá", "don gia", "selling price", "price adjustment", "doanh thu", "revenue"],
    "investment_cost_change": ["tổng mức đầu tư", "tong muc dau tu", "tmđt", "tmdt", "chi phí đầu tư", "chi phi dau tu", "investment cost", "capex", "total investment"],
    "selling_cost_rate": ["chi phí bán hàng", "chi phi ban hang", "hoa hồng", "hoa hong", "môi giới", "moi gioi", "selling cost", "brokerage"],
    "marketing_rate": ["marketing", "truyền thông", "truyen thong", "chi phí marketing", "chi phi marketing"],
    "admin_cost_rate": ["chi phí quản lý", "chi phi quan ly", "g&a", "admin cost", "management cost"],
    "sales_volume_change": ["sản lượng", "san luong", "số căn", "so can", "sales volume", "units sold"],
    "debt_financing_ratio": ["tỷ lệ vay", "ty le vay", "vốn vay", "von vay", "debt ratio", "loan ratio"],
}

OUTPUT_KEYWORDS: dict[str, list[str]] = {
    "project_npv": ["npv dự án", "npv du an", "project npv", "npv"],
    "project_irr": ["irr dự án", "irr du an", "project irr", "irr"],
    "equity_npv": ["npv chủ đầu tư", "npv chu dau tu", "equity npv", "investor npv"],
    "equity_irr": ["irr chủ đầu tư", "irr chu dau tu", "equity irr", "investor irr"],
    "roi": ["roi"],
    "profit_after_tax": ["lợi nhuận sau thuế", "loi nhuan sau thue", "profit after tax", "lnst"],
    "total_investment": ["tổng mức đầu tư", "tong muc dau tu", "tmđt", "tmdt", "total investment"],
    "bank_loan": ["vay ngân hàng", "vay ngan hang", "bank loan", "loan amount"],
    "total_revenue": ["tổng doanh thu", "tong doanh thu", "total revenue"],
    "revenue": ["doanh thu", "revenue"],
}

DEFAULT_OUTPUT_ORDER = [
    "profit_after_tax", "project_npv", "project_irr", "equity_npv", "equity_irr", "roi", "total_investment", "bank_loan", "total_revenue", "revenue"
]

WORKBOOK_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"


def is_formula(value: Any) -> bool:
    return isinstance(value, str) and value.startswith("=")


def compact_value(value: Any, max_len: int = 90) -> str:
    if value is None:
        return ""
    text = str(value)
    return text if len(text) <= max_len else text[: max_len - 3] + "..."


def matches(text: Any, phrases: list[str]) -> bool:
    if not text:
        return False
    norm = normalize_text(str(text))
    return any(normalize_text(p) in norm for p in phrases)


def guess_param(label: str, keyword_map: dict[str, list[str]]) -> str:
    for key, phrases in keyword_map.items():
        if matches(label, phrases):
            return key
    return ""


def nearby_label(ws: Any, row: int, col: int) -> str:
    candidates: list[str] = []
    # Prefer text immediately to the left, then further left, then above.
    for offset in range(1, 5):
        c = col - offset
        if c >= 1:
            v = ws.cell(row=row, column=c).value
            if isinstance(v, str) and v.strip():
                candidates.append(v.strip())
                break
    for offset in range(1, 4):
        r = row - offset
        if r >= 1:
            v = ws.cell(row=r, column=col).value
            if isinstance(v, str) and v.strip():
                candidates.append(v.strip())
                break
    # Also try right side for sheets where labels appear after values.
    for offset in range(1, 3):
        c = col + offset
        if c <= ws.max_column:
            v = ws.cell(row=row, column=c).value
            if isinstance(v, str) and v.strip():
                candidates.append(v.strip())
                break
    return " | ".join(dict.fromkeys(candidates))


def type_guess(value: Any, label: str, param: str = "") -> str:
    norm = normalize_text(str(label or ""))
    if param.endswith("_irr") or param == "roi" or "irr" in norm or "roi" in norm:
        return "percent"
    if "vat" in norm or "tndn" in norm or "lai" in norm or "rate" in norm or "ty le" in norm:
        return "percent"
    if isinstance(value, (int, float)):
        if -1.5 <= float(value) <= 1.5 and ("%" in norm or "rate" in norm or "lai" in norm or "ty le" in norm):
            return "percent"
        if abs(float(value)) > 1000:
            return "currency"
    if "npv" in norm or "doanh thu" in norm or "chi phi" in norm or "vnd" in norm or "tien" in norm:
        return "currency"
    return "number"


def default_min_max(param: str, t: str) -> tuple[Any, Any]:
    if param == "loan_interest_rate":
        return 0.0, 0.25
    if param == "vat_rate":
        return 0.0, 0.20
    if param == "cit_rate":
        return 0.0, 0.35
    if param in {"selling_price_change", "sales_volume_change"}:
        return -0.50, 1.00
    if param == "investment_cost_change":
        return -0.30, 1.00
    if t == "percent":
        return -1.0, 1.0
    return None, None


def infer_unit(t: str) -> str:
    if t == "percent":
        return "decimal"
    if t == "currency":
        return "VND"
    return ""


def confidence_for(role: str, param: str, label: str, formula: bool) -> str:
    if not param:
        return "low"
    if role == "input":
        return "high" if not formula else "low"
    if role == "output":
        return "high" if formula else "medium"
    return "medium"


def aliases_for(param: str, role: str) -> list[str]:
    source = INPUT_KEYWORDS if role == "input" else OUTPUT_KEYWORDS
    return source.get(param, [])


def sanitize_workbook_defined_names(excel_path: Path) -> list[str]:
    """Remove broken workbook defined names that make openpyxl refuse to load.

    Some Excel files contain stale print-area/print-title metadata such as
    `_xlnm.Print_Titles = #N/A`. Excel may still open the file, but openpyxl
    fails before we can scan any sheet. The uploaded workbook is already a copy,
    so it is safe to remove only these invalid metadata entries.
    """
    warnings_out: list[str] = []
    if excel_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return warnings_out

    with zipfile.ZipFile(excel_path, "r") as zin:
        try:
            workbook_xml = zin.read("xl/workbook.xml")
        except KeyError:
            return warnings_out

        root = ET.fromstring(workbook_xml)
        ns = {"x": WORKBOOK_NS}
        defined_names = root.find("x:definedNames", ns)
        if defined_names is None:
            return warnings_out

        removed = 0
        for defined_name in list(defined_names):
            name = str(defined_name.attrib.get("name", ""))
            value = (defined_name.text or "").strip()
            is_print_metadata = name in {"_xlnm.Print_Titles", "_xlnm.Print_Area"}
            is_invalid_ref = value in {"#N/A", "#REF!", "#VALUE!"} or value.startswith("#")
            if is_invalid_ref:
                defined_names.remove(defined_name)
                removed += 1
                scope = defined_name.attrib.get("localSheetId")
                scope_text = f" localSheetId={scope}" if scope is not None else ""
                reason = "print metadata" if is_print_metadata else "broken reference"
                warnings_out.append(f"Removed invalid workbook defined name ({reason}): {name}{scope_text}={value}")

        if removed == 0:
            return warnings_out
        if len(list(defined_names)) == 0:
            root.remove(defined_names)

        ET.register_namespace("", WORKBOOK_NS)
        fixed_xml = ET.tostring(root, encoding="utf-8", xml_declaration=True)

        with tempfile.NamedTemporaryFile(delete=False, suffix=excel_path.suffix) as tmp:
            tmp_path = Path(tmp.name)

        try:
            with zipfile.ZipFile(tmp_path, "w") as zout:
                for item in zin.infolist():
                    if item.filename == "xl/workbook.xml":
                        zout.writestr(item, fixed_xml)
                    else:
                        zout.writestr(item, zin.read(item.filename))
            shutil.move(str(tmp_path), excel_path)
        finally:
            if tmp_path.exists():
                tmp_path.unlink(missing_ok=True)

    return warnings_out


def scan_workbook_for_mapping(excel_path: Path) -> dict[str, Any]:
    preflight_warnings = sanitize_workbook_defined_names(excel_path)
    with warnings.catch_warnings(record=True) as caught_warnings:
        warnings.simplefilter("always")
        wb = load_workbook(excel_path, data_only=False, read_only=False)
    candidates: list[dict[str, Any]] = []
    formula_errors: list[dict[str, Any]] = []
    scan_warnings: list[str] = preflight_warnings + [str(w.message) for w in caught_warnings]
    sheets = [{"title": ws.title, "max_row": ws.max_row, "max_column": ws.max_column} for ws in wb.worksheets]

    seen: set[tuple[str, str]] = set()
    try:
        for ws in wb.worksheets:
            # Cap huge sheets for UI responsiveness while still covering normal financial summary sheets.
            for row in ws.iter_rows(max_row=min(ws.max_row, 500), max_col=min(ws.max_column, 80)):
                for cell in row:
                    try:
                        value = cell.value
                        if value is None:
                            continue
                        formula = is_formula(value)
                        label = nearby_label(ws, cell.row, cell.column)
                        value_is_numeric = isinstance(value, (int, float))
                        value_is_formula_or_numeric = value_is_numeric or formula
                        if isinstance(value, str) and any(token in value for token in ["#REF!", "#VALUE!", "#DIV/0!", "#N/A"]):
                            formula_errors.append({
                                "sheet": ws.title,
                                "cell": cell.coordinate,
                                "value": compact_value(value),
                                "nearby_label": label,
                            })

                        if not label:
                            continue

                        # Inputs: hardcoded numeric values near input-ish labels.
                        if value_is_numeric and not formula:
                            param = guess_param(label, INPUT_KEYWORDS)
                            if param:
                                key = ("input", param)
                                t = type_guess(value, label, param)
                                min_v, max_v = default_min_max(param, t)
                                candidate = {
                                    "id": f"input::{param}::{ws.title}!{cell.coordinate}",
                                    "decision": "Approve",
                                    "role": "input",
                                    "parameter_key": param,
                                    "friendly_name": FRIENDLY_NAMES.get(param, param),
                                    "sheet": ws.title,
                                    "cell": cell.coordinate,
                                    "current_value": compact_value(value),
                                    "formula": False,
                                    "type": t,
                                    "unit": infer_unit(t),
                                    "min": min_v,
                                    "max": max_v,
                                    "editable": True,
                                    "confidence": confidence_for("input", param, label, False),
                                    "nearby_label": label,
                                    "description": f"Candidate input found near label: {label}",
                                    "aliases": aliases_for(param, "input"),
                                    "target_mode": "cell",
                                    "notes": "",
                                }
                                if key not in seen:
                                    seen.add(key)
                                    candidates.append(candidate)

                        # Outputs: numeric or formula cells near output-ish labels.
                        if value_is_formula_or_numeric:
                            param = guess_param(label, OUTPUT_KEYWORDS)
                            if param:
                                key = ("output", param)
                                t = type_guess(value, label, param)
                                candidate = {
                                    "id": f"output::{param}::{ws.title}!{cell.coordinate}",
                                    "decision": "Approve" if confidence_for("output", param, label, formula) == "high" else "Review",
                                    "role": "output",
                                    "parameter_key": param,
                                    "friendly_name": FRIENDLY_NAMES.get(param, param),
                                    "sheet": ws.title,
                                    "cell": cell.coordinate,
                                    "current_value": compact_value(value),
                                    "formula": bool(formula),
                                    "type": t,
                                    "unit": infer_unit(t),
                                    "min": None,
                                    "max": None,
                                    "editable": False,
                                    "confidence": confidence_for("output", param, label, formula),
                                    "nearby_label": label,
                                    "description": f"Candidate output found near label: {label}",
                                    "aliases": aliases_for(param, "output"),
                                    "target_mode": "cell",
                                    "notes": "",
                                }
                                if key not in seen:
                                    seen.add(key)
                                    candidates.append(candidate)
                    except Exception as exc:
                        scan_warnings.append(f"Skipped {ws.title}!{cell.coordinate}: {exc}")
    finally:
        wb.close()

    # Prioritize approved/high confidence first, then inputs, then outputs.
    score = {"high": 0, "medium": 1, "low": 2}
    candidates.sort(key=lambda x: (score.get(str(x.get("confidence")), 9), 0 if x.get("role") == "input" else 1, x.get("parameter_key", "")))
    return {
        "excel": str(excel_path),
        "workbook_name": excel_path.name,
        "suggested_profile": safe_filename(excel_path.stem),
        "sheets": sheets,
        "candidates": candidates,
        "formula_errors": formula_errors,
        "scan_warnings": scan_warnings[:100],
        "summary": {
            "candidate_count": len(candidates),
            "formula_error_count": len(formula_errors),
            "scan_warning_count": len(scan_warnings),
            "sheet_count": len(sheets),
        },
    }


def split_aliases(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, list):
        raw = value
    else:
        raw = str(value).replace(";", "|").replace(",", "|").split("|")
    return sorted({str(v).strip() for v in raw if str(v).strip()}, key=lambda x: x.lower())


def as_bool(value: Any, default: bool = True) -> bool:
    text = str(value if value is not None else "").strip().lower()
    if text in {"yes", "true", "1", "approve", "approved", "y"}:
        return True
    if text in {"no", "false", "0", "reject", "rejected", "n"}:
        return False
    return default


def clean_number(value: Any) -> Any:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return value
    try:
        return float(str(value).strip())
    except ValueError:
        return None


def target_fields(item: dict[str, Any]) -> dict[str, Any]:
    sheet = str(item.get("correct_sheet") or item.get("sheet") or "").strip()
    target = str(item.get("correct_cell") or item.get("cell") or "").strip()
    mode = str(item.get("target_mode") or "cell").strip().lower()
    if not sheet or not target:
        raise ValueError(f"Approved mapping lacks sheet/cell: {item}")
    fields: dict[str, Any] = {"sheet": sheet}
    if mode == "range" or ":" in target:
        fields["range"] = target
    elif mode == "cells" or "," in target:
        fields["cells"] = [c.strip() for c in target.split(",") if c.strip()]
    else:
        fields["cell"] = target
    return fields


def build_model_map_from_candidates(profile_name: str, candidates: list[dict[str, Any]], currency: str = "VND") -> dict[str, Any]:
    inputs: dict[str, Any] = {}
    outputs: dict[str, Any] = {}
    approved = 0
    skipped = 0

    for item in candidates:
        decision = str(item.get("decision") or "").strip().lower()
        if decision not in {"approve", "approved", "yes"}:
            skipped += 1
            continue
        role = str(item.get("role") or "").strip().lower()
        key = str(item.get("parameter_key") or "").strip()
        if not key:
            skipped += 1
            continue
        fields = target_fields(item)
        t = str(item.get("type") or "number").strip() or "number"
        unit = str(item.get("unit") or "").strip() or None
        desc = str(item.get("description") or item.get("friendly_name") or key).strip()
        aliases = split_aliases(item.get("aliases"))
        if role == "input":
            meta = {
                **fields,
                "type": t,
                "editable": as_bool(item.get("editable"), True),
                "min": clean_number(item.get("min")),
                "max": clean_number(item.get("max")),
                "unit": unit,
                "description": desc,
                "aliases": aliases,
            }
            inputs[key] = {k: v for k, v in meta.items() if v is not None and v != []}
            approved += 1
        elif role == "output":
            if "range" in fields or "cells" in fields:
                raise ValueError(f"Output {key} must map to a single cell, not a range/list: {fields}")
            meta = {**fields, "type": t, "unit": unit, "description": desc, "aliases": aliases}
            outputs[key] = {k: v for k, v in meta.items() if v is not None and v != []}
            approved += 1
        else:
            skipped += 1

    default_outputs = [k for k in DEFAULT_OUTPUT_ORDER if k in outputs] or list(outputs.keys())[:8]
    return {
        "metadata": {
            "name": profile_name,
            "version": "ui-onboarded-0.2",
            "currency": currency,
            "note": "Generated from web onboarding wizard. Validate before production use.",
            "review_stats": {"approved_rows": approved, "skipped_rows": skipped},
        },
        "settings": {"allow_formula_overwrite": False, "default_outputs": default_outputs},
        "inputs": inputs,
        "outputs": outputs,
    }


def input_targets(meta: dict[str, Any]) -> list[str]:
    if meta.get("cell"):
        return [str(meta["cell"])]
    if meta.get("cells"):
        return [str(c) for c in meta["cells"]]
    if meta.get("range"):
        min_col, min_row, max_col, max_row = range_boundaries(str(meta["range"]))
        return [f"{get_column_letter(c)}{r}" for r in range(min_row, max_row + 1) for c in range(min_col, max_col + 1)]
    return []


def validate_model_map_against_workbook(model_map: dict[str, Any], excel_path: Path) -> dict[str, Any]:
    sanitize_workbook_defined_names(excel_path)
    wb = load_workbook(excel_path, data_only=False)
    errors: list[str] = []
    warnings: list[str] = []

    for key, meta in model_map.get("inputs", {}).items():
        sheet = meta.get("sheet")
        if sheet not in wb.sheetnames:
            errors.append(f"Input {key}: sheet not found: {sheet}")
            continue
        targets = input_targets(meta)
        if not targets:
            errors.append(f"Input {key}: must define cell, cells, or range")
            continue
        ws = wb[sheet]
        for target in targets:
            try:
                value = ws[target].value
            except Exception as exc:
                errors.append(f"Input {key}: invalid cell {sheet}!{target}: {exc}")
                continue
            if is_formula(value) and not meta.get("allow_formula_overwrite", False):
                errors.append(f"Input {key}: selected cell {sheet}!{target} contains a formula. Choose a real input cell.")
            if value is None:
                warnings.append(f"Input {key}: selected cell is blank: {sheet}!{target}")

    for key, meta in model_map.get("outputs", {}).items():
        sheet = meta.get("sheet")
        cell = meta.get("cell")
        if sheet not in wb.sheetnames:
            errors.append(f"Output {key}: sheet not found: {sheet}")
            continue
        if not cell:
            errors.append(f"Output {key}: must define a single cell")
            continue
        try:
            _ = wb[sheet][cell].value
        except Exception as exc:
            errors.append(f"Output {key}: invalid cell {sheet}!{cell}: {exc}")

    return {"ok": not errors, "errors": errors, "warnings": warnings}


def save_profile_yaml(base_dir: Path, profile_name: str, model_map: dict[str, Any], overwrite: bool = True) -> Path:
    safe_profile = safe_filename(profile_name)
    profile_dir = base_dir / "config" / "profiles" / safe_profile
    profile_dir.mkdir(parents=True, exist_ok=True)
    out_path = profile_dir / "model_map.yaml"
    if out_path.exists() and not overwrite:
        raise FileExistsError(f"Profile already exists: {safe_profile}")
    out_path.write_text(yaml.safe_dump(model_map, allow_unicode=True, sort_keys=False), encoding="utf-8")
    readme = profile_dir / "README.md"
    readme.write_text(
        f"# Profile: {safe_profile}\n\nGenerated by the web onboarding wizard.\n\nValidate before production use.\n",
        encoding="utf-8",
    )
    return out_path
