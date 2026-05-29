from __future__ import annotations

from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, range_boundaries

from .utils import normalize_text


def _tokens(text: str) -> set[str]:
    return {token for token in normalize_text(text).replace("/", " ").split() if token}

from .utils import safe_filename


def _copy_cell(src: Any, dst: Any, value: Any) -> None:
    dst.value = value
    if src.has_style:
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)
    if src.hyperlink:
        dst._hyperlink = copy(src.hyperlink)
    if src.comment:
        dst.comment = copy(src.comment)


def export_sensitivity_range(
    source_path: Path,
    sheet_name: str,
    range_address: str,
    output_dir: Path,
    output_name: str | None = None,
) -> Path:
    if not source_path.exists():
        raise FileNotFoundError(f"Excel source file not found: {source_path}")

    try:
        from .onboarding import sanitize_workbook_defined_names

        sanitize_workbook_defined_names(source_path)
    except Exception:
        pass

    output_dir.mkdir(parents=True, exist_ok=True)
    source_wb = load_workbook(source_path, data_only=False, read_only=False)
    values_wb = load_workbook(source_path, data_only=True, read_only=False)
    try:
        if sheet_name not in source_wb.sheetnames:
            raise ValueError(f"Sheet not found: {sheet_name}")
        if sheet_name not in values_wb.sheetnames:
            raise ValueError(f"Sheet not found in cached values: {sheet_name}")

        min_col, min_row, max_col, max_row = range_boundaries(range_address)
        source_ws = source_wb[sheet_name]
        values_ws = values_wb[sheet_name]

        out_wb = Workbook()
        out_ws = out_wb.active
        out_ws.title = "Sensitivity"
        out_ws.freeze_panes = "A2"

        for row in range(min_row, max_row + 1):
            out_row = row - min_row + 1
            if source_ws.row_dimensions[row].height:
                out_ws.row_dimensions[out_row].height = source_ws.row_dimensions[row].height
            for col in range(min_col, max_col + 1):
                out_col = col - min_col + 1
                src = source_ws.cell(row=row, column=col)
                val = values_ws.cell(row=row, column=col).value
                dst = out_ws.cell(row=out_row, column=out_col)
                _copy_cell(src, dst, val)

        for col in range(min_col, max_col + 1):
            src_letter = get_column_letter(col)
            out_letter = get_column_letter(col - min_col + 1)
            width = source_ws.column_dimensions[src_letter].width
            if width:
                out_ws.column_dimensions[out_letter].width = width

        for merged in source_ws.merged_cells.ranges:
            if (
                merged.min_col >= min_col
                and merged.max_col <= max_col
                and merged.min_row >= min_row
                and merged.max_row <= max_row
            ):
                out_ws.merge_cells(
                    start_row=merged.min_row - min_row + 1,
                    start_column=merged.min_col - min_col + 1,
                    end_row=merged.max_row - min_row + 1,
                    end_column=merged.max_col - min_col + 1,
                )

        stem = safe_filename(output_name or f"sensitivity_{sheet_name}_{range_address}")
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = output_dir / f"{stamp}_{stem}.xlsx"
        out_wb.save(out_path)
        return out_path
    finally:
        source_wb.close()
        values_wb.close()


def _important_terms(prompt: str) -> list[str]:
    stopwords = {
        "cho",
        "toi",
        "to",
        "xuat",
        "export",
        "bang",
        "table",
        "do",
        "nhay",
        "sensitivity",
        "cua",
        "hien",
        "thi",
    }
    words = [w for w in normalize_text(prompt).replace("/", " ").split() if w and w not in stopwords]
    return words or [normalize_text(prompt)]


def _matched_terms(text: Any, terms: list[str]) -> set[str]:
    if not isinstance(text, str):
        return set()
    text_norm = normalize_text(text)
    text_tokens = _tokens(text)
    matched: set[str] = set()
    for term in terms:
        if " " in term:
            if term in text_norm:
                matched.add(term)
        elif term in text_tokens:
            matched.add(term)
    return matched


def _cell_score(value: Any, terms: list[str], prompt_norm: str) -> int:
    if not isinstance(value, str) or not value.strip():
        return 0
    text = normalize_text(value)
    score = 0
    if prompt_norm and prompt_norm in text:
        score += 100
    if "do nhay" in text or "sensitivity" in text:
        score += 40
    matched = _matched_terms(value, terms)
    for term in matched:
        score += 15
    if len(matched) == len(terms):
        score += 45
    return score


def _has_content(ws: Any, row: int, min_col: int, max_col: int) -> bool:
    for col in range(min_col, max_col + 1):
        if ws.cell(row=row, column=col).value not in (None, ""):
            return True
    return False


def _is_next_sensitivity_title(value: Any) -> bool:
    if not isinstance(value, str):
        return False
    text = normalize_text(value)
    return "do nhay" in text or "sensitivity" in text


def _detect_table_range(ws: Any, anchor_row: int, anchor_col: int) -> str:
    min_col = max(1, anchor_col)
    min_row = max(1, anchor_row)
    search_max_col = min(ws.max_column, min_col + 28)
    search_max_row = min(ws.max_row, min_row + 70)

    max_row = min_row
    blank_streak = 0
    for row in range(min_row, search_max_row + 1):
        if row > min_row and any(_is_next_sensitivity_title(ws.cell(row=row, column=col).value) for col in range(1, search_max_col + 1)):
            break
        if _has_content(ws, row, min_col, search_max_col):
            max_row = row
            blank_streak = 0
        else:
            blank_streak += 1
            if blank_streak >= 3 and row > min_row + 4:
                break

    max_col = min_col
    blank_streak = 0
    for col in range(min_col, search_max_col + 1):
        has_col_content = any(ws.cell(row=row, column=col).value not in (None, "") for row in range(min_row, max_row + 1))
        if has_col_content:
            max_col = col
            blank_streak = 0
        else:
            blank_streak += 1
            if blank_streak >= 3 and col > min_col + 4:
                break

    return f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"


def find_sensitivity_range(source_path: Path, prompt: str) -> dict[str, Any]:
    if not source_path.exists():
        raise FileNotFoundError(f"Excel source file not found: {source_path}")

    try:
        from .onboarding import sanitize_workbook_defined_names

        sanitize_workbook_defined_names(source_path)
    except Exception:
        pass

    prompt_norm = normalize_text(prompt)
    terms = _important_terms(prompt)
    wb = load_workbook(source_path, data_only=False, read_only=False)
    try:
        best: tuple[int, str, int, int, str] | None = None
        candidates: list[dict[str, Any]] = []
        for ws in wb.worksheets:
            for row in ws.iter_rows(max_row=min(ws.max_row, 600), max_col=min(ws.max_column, 120)):
                for cell in row:
                    score = _cell_score(cell.value, terms, prompt_norm)
                    if score <= 0:
                        continue
                    matched = sorted(_matched_terms(cell.value, terms))
                    candidates.append({
                        "sheet": ws.title,
                        "cell": cell.coordinate,
                        "label": str(cell.value),
                        "score": score,
                        "matched_terms": matched,
                    })
                    candidate = (score, ws.title, cell.row, cell.column, str(cell.value))
                    if best is None or candidate[0] > best[0]:
                        best = candidate

        if best is None or best[0] < 45:
            raise ValueError("Could not find a sensitivity table matching that prompt. Try specifying sheet/range.")

        _, sheet_name, row, col, label = best
        required_terms = set(terms)
        matched_terms = _matched_terms(label, terms)
        if not required_terms.issubset(matched_terms):
            top = sorted(candidates, key=lambda x: x["score"], reverse=True)[:5]
            suggestions = "; ".join(f"{c['sheet']}!{c['cell']} = {c['label']}" for c in top)
            raise ValueError(
                "Could not find an exact sensitivity table for that prompt. "
                f"Closest matches: {suggestions}. "
                "Try a more specific prompt matching the table title, or use sheet/range export."
            )

        ws = wb[sheet_name]
        range_address = _detect_table_range(ws, row, col)
        return {
            "sheet": sheet_name,
            "range_address": range_address,
            "matched_label": label,
            "score": best[0],
            "matched_terms": sorted(matched_terms),
        }
    finally:
        wb.close()


def export_sensitivity_by_prompt(
    source_path: Path,
    prompt: str,
    output_dir: Path,
    output_name: str | None = None,
) -> tuple[Path, dict[str, Any]]:
    detected = find_sensitivity_range(source_path, prompt)
    out_path = export_sensitivity_range(
        source_path=source_path,
        sheet_name=detected["sheet"],
        range_address=detected["range_address"],
        output_dir=output_dir,
        output_name=output_name or prompt,
    )
    return out_path, detected
