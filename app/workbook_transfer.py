from __future__ import annotations

import shutil
import zipfile
import re
from datetime import datetime
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET
from xml.sax.saxutils import escape

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries

from .utils import safe_filename


SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm"}
MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PKG_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"

ET.register_namespace("", MAIN_NS)
ET.register_namespace("r", REL_NS)


def _is_formula(value: Any) -> bool:
    return isinstance(value, str) and value.startswith("=")


def _display_value(value: Any) -> Any:
    if value is None:
        return None
    return value


def _serializable(value: Any) -> Any:
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return value


def _workbook_kwargs(path: Path) -> dict[str, Any]:
    return {"keep_vba": path.suffix.lower() == ".xlsm"}


def _validate_workbook_path(path: Path, label: str) -> None:
    if not path.exists():
        raise FileNotFoundError(f"{label} workbook not found: {path}")
    if path.suffix.lower() not in SUPPORTED_EXTENSIONS:
        raise ValueError(f"{label} workbook must be .xlsx or .xlsm: {path.name}")


def _matching_sheets(original_wb: Any, edited_wb: Any, sheet_name: str | None = None) -> list[str]:
    if sheet_name:
        if sheet_name not in original_wb.sheetnames:
            raise ValueError(f"Sheet not found in original workbook: {sheet_name}")
        if sheet_name not in edited_wb.sheetnames:
            raise ValueError(f"Sheet not found in edited workbook: {sheet_name}")
        return [sheet_name]
    return [sheet for sheet in edited_wb.sheetnames if sheet in original_wb.sheetnames]


def _expand_range(ref: str) -> list[str]:
    min_col, min_row, max_col, max_row = range_boundaries(ref)
    cells: list[str] = []
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cells.append(f"{get_column_letter(col)}{row}")
    return cells


def _shared_formula_cells_by_sheet(xlsx_path: Path) -> dict[str, dict[str, set[str]]]:
    try:
        sheet_paths = _sheet_xml_paths_by_name(xlsx_path)
    except Exception:
        return {}

    out: dict[str, dict[str, set[str]]] = {}
    with zipfile.ZipFile(xlsx_path, "r") as zf:
        for sheet_name, part in sheet_paths.items():
            if part not in zf.namelist():
                continue
            root = ET.fromstring(zf.read(part))
            si_ranges: dict[str, set[str]] = {}
            si_seen: dict[str, set[str]] = {}
            for cell in root.findall(f".//{{{MAIN_NS}}}c"):
                ref = cell.attrib.get("r")
                formula = cell.find(f"{{{MAIN_NS}}}f")
                if not ref or formula is None or formula.attrib.get("t") != "shared":
                    continue
                si = formula.attrib.get("si")
                if not si:
                    continue
                si_seen.setdefault(si, set()).add(ref)
                if formula.attrib.get("ref"):
                    si_ranges[si] = set(_expand_range(formula.attrib["ref"]))

            cell_to_group: dict[str, set[str]] = {}
            for si, cells in si_seen.items():
                group = si_ranges.get(si, cells)
                for cell_ref in group:
                    cell_to_group[cell_ref] = group
            out[sheet_name] = cell_to_group
    return out


def plan_workbook_value_transfer(
    original_path: Path,
    edited_path: Path,
    sheet_name: str | None = None,
    include_formula_cells: bool = False,
    include_blank_cells: bool = False,
    max_preview: int = 100,
) -> dict[str, Any]:
    """Compare two workbooks and return cells that can be copied safely.

    The edited workbook is treated as the source of displayed values. The
    original workbook is never modified by this function. Formula cells in the
    original workbook are skipped by default because replacing them with values
    can break the model.
    """
    _validate_workbook_path(original_path, "Original")
    _validate_workbook_path(edited_path, "Edited")

    original_wb = load_workbook(original_path, data_only=False, read_only=False, **_workbook_kwargs(original_path))
    original_values_wb = load_workbook(original_path, data_only=True, read_only=False, **_workbook_kwargs(original_path))
    edited_formula_wb = load_workbook(edited_path, data_only=False, read_only=False, **_workbook_kwargs(edited_path))
    edited_values_wb = load_workbook(edited_path, data_only=True, read_only=False, **_workbook_kwargs(edited_path))
    try:
        sheets = _matching_sheets(original_wb, edited_formula_wb, sheet_name)
        if not sheets:
            raise ValueError("No matching sheet names found between original and edited workbook.")

        changes_by_key: dict[tuple[str, str], dict[str, Any]] = {}
        shared_cells_by_sheet = _shared_formula_cells_by_sheet(original_path) if include_formula_cells else {}
        shared_groups_to_snapshot: dict[str, set[frozenset[str]]] = {}
        skipped_formula = 0
        skipped_blank = 0
        unchanged = 0
        scanned_cells = 0

        for sheet in sheets:
            original_ws = original_wb[sheet]
            original_values_ws = original_values_wb[sheet]
            edited_formula_ws = edited_formula_wb[sheet]
            edited_values_ws = edited_values_wb[sheet]
            max_row = edited_formula_ws.max_row
            max_col = edited_formula_ws.max_column

            for row in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    scanned_cells += 1
                    cell_ref = f"{get_column_letter(col)}{row}"
                    target_cell = original_ws.cell(row=row, column=col)
                    edited_formula_value = edited_formula_ws.cell(row=row, column=col).value
                    old_display_value = _display_value(original_values_ws.cell(row=row, column=col).value)
                    new_display_value = _display_value(edited_values_ws.cell(row=row, column=col).value)

                    if new_display_value in (None, "") and not include_blank_cells:
                        skipped_blank += 1
                        continue
                    if old_display_value == new_display_value:
                        unchanged += 1
                        continue

                    target_is_formula = _is_formula(target_cell.value)
                    source_is_formula = _is_formula(edited_formula_value)
                    if target_is_formula and not include_formula_cells:
                        skipped_formula += 1
                        continue

                    item = {
                        "sheet": sheet,
                        "cell": cell_ref,
                        "old_value": _serializable(old_display_value),
                        "new_value": _serializable(new_display_value),
                        "target_is_formula": target_is_formula,
                        "source_is_formula": source_is_formula,
                    }
                    changes_by_key[(sheet, cell_ref)] = item
                    group = shared_cells_by_sheet.get(sheet, {}).get(cell_ref)
                    if target_is_formula and group:
                        shared_groups_to_snapshot.setdefault(sheet, set()).add(frozenset(group))

        for sheet, groups in shared_groups_to_snapshot.items():
            original_ws = original_wb[sheet]
            original_values_ws = original_values_wb[sheet]
            edited_formula_ws = edited_formula_wb[sheet]
            edited_values_ws = edited_values_wb[sheet]
            for group in groups:
                for cell_ref in group:
                    new_display_value = _display_value(edited_values_ws[cell_ref].value)
                    if new_display_value in (None, "") and not include_blank_cells:
                        continue
                    old_display_value = _display_value(original_values_ws[cell_ref].value)
                    edited_formula_value = edited_formula_ws[cell_ref].value
                    target_is_formula = _is_formula(original_ws[cell_ref].value)
                    source_is_formula = _is_formula(edited_formula_value)
                    changes_by_key[(sheet, cell_ref)] = {
                        "sheet": sheet,
                        "cell": cell_ref,
                        "old_value": _serializable(old_display_value),
                        "new_value": _serializable(new_display_value),
                        "target_is_formula": target_is_formula,
                        "source_is_formula": source_is_formula,
                        "shared_formula_snapshot": True,
                    }

        changes = sorted(changes_by_key.values(), key=lambda item: (item["sheet"], item["cell"]))

        shared_snapshot_count = sum(1 for item in changes if item.get("shared_formula_snapshot"))
        warnings = _warnings_for_transfer(len(changes), skipped_formula, include_formula_cells)
        if shared_snapshot_count:
            warnings.append(
                f"Expanded {shared_snapshot_count} cells from shared formula groups to prevent Excel repair."
            )

        return {
            "ok": True,
            "original_file": original_path.name,
            "edited_file": edited_path.name,
            "sheets": sheets,
            "preserved_original_sheets": original_wb.sheetnames,
            "output_keeps_unmodified_original_content": True,
            "scanned_cells": scanned_cells,
            "change_count": len(changes),
            "skipped_formula_count": skipped_formula,
            "skipped_blank_count": skipped_blank,
            "unchanged_count": unchanged,
            "include_formula_cells": include_formula_cells,
            "include_blank_cells": include_blank_cells,
            "shared_formula_snapshot_count": shared_snapshot_count,
            "preview_limit": max_preview,
            "changes_preview": changes[:max_preview],
            "preview_truncated": len(changes) > max_preview,
            "warnings": warnings,
        }
    finally:
        original_wb.close()
        original_values_wb.close()
        edited_formula_wb.close()
        edited_values_wb.close()


def _warnings_for_transfer(change_count: int, skipped_formula: int, include_formula_cells: bool) -> list[str]:
    warnings: list[str] = []
    if skipped_formula and not include_formula_cells:
        warnings.append(
            f"Skipped {skipped_formula} formula cells. This protects formulas in the original workbook."
        )
    if include_formula_cells:
        warnings.append(
            "Formula cells will be replaced with fixed displayed values in the exported copy. "
            "Use only when you want a static snapshot."
        )
    if change_count == 0:
        warnings.append("No writable value changes were found with the current safeguards.")
    return warnings


def _sheet_xml_paths_by_name(xlsx_path: Path) -> dict[str, str]:
    with zipfile.ZipFile(xlsx_path, "r") as zf:
        workbook_xml = ET.fromstring(zf.read("xl/workbook.xml"))
        rels_xml = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))

    rel_targets: dict[str, str] = {}
    for rel in rels_xml.findall(f"{{{PKG_REL_NS}}}Relationship"):
        rel_id = rel.attrib.get("Id")
        target = rel.attrib.get("Target", "")
        if not rel_id or not target:
            continue
        if target.startswith("/"):
            path = target.lstrip("/")
        elif target.startswith("xl/"):
            path = target
        else:
            path = f"xl/{target}"
        rel_targets[rel_id] = path.replace("\\", "/")

    out: dict[str, str] = {}
    for sheet in workbook_xml.findall(f".//{{{MAIN_NS}}}sheet"):
        name = sheet.attrib.get("name")
        rel_id = sheet.attrib.get(f"{{{REL_NS}}}id")
        if name and rel_id and rel_id in rel_targets:
            out[name] = rel_targets[rel_id]
    return out


def _remove_cell_payload(cell: ET.Element) -> None:
    for child in list(cell):
        if child.tag in {f"{{{MAIN_NS}}}f", f"{{{MAIN_NS}}}v", f"{{{MAIN_NS}}}is"}:
            cell.remove(child)
    for attr in ["t", "cm", "vm"]:
        cell.attrib.pop(attr, None)


def _set_cell_value(cell: ET.Element, value: Any) -> None:
    _remove_cell_payload(cell)
    if value in (None, ""):
        return
    if isinstance(value, bool):
        cell.set("t", "b")
        v = ET.SubElement(cell, f"{{{MAIN_NS}}}v")
        v.text = "1" if value else "0"
        return
    if isinstance(value, (int, float)):
        v = ET.SubElement(cell, f"{{{MAIN_NS}}}v")
        v.text = repr(value)
        return
    cell.set("t", "inlineStr")
    inline = ET.SubElement(cell, f"{{{MAIN_NS}}}is")
    text = ET.SubElement(inline, f"{{{MAIN_NS}}}t")
    text.text = str(value)


def _cell_value_xml(value: Any) -> tuple[str, str]:
    if value in (None, ""):
        return "", ""
    if isinstance(value, bool):
        return ' t="b"', f"<v>{'1' if value else '0'}</v>"
    if isinstance(value, (int, float)):
        return "", f"<v>{repr(value)}</v>"
    return ' t="inlineStr"', f"<is><t>{escape(str(value))}</t></is>"


def _strip_cell_value_attrs(attrs: str) -> str:
    attrs = re.sub(r'\s(?:t|cm|vm)="[^"]*"', "", attrs)
    return attrs


def _patch_cell_fragment(fragment: str, value: Any) -> str:
    m = re.match(r"(?s)(<c\b)([^>]*)(/>|>.*?</c>)", fragment)
    if not m:
        return fragment
    start, attrs, rest = m.groups()
    attrs = _strip_cell_value_attrs(attrs)
    type_attr, payload = _cell_value_xml(value)
    if type_attr:
        attrs += type_attr
    return f"{start}{attrs}>{payload}</c>"


def _patch_sheet_xml(xml_bytes: bytes, changes: list[dict[str, Any]]) -> bytes:
    # Patch only individual <c r="...">...</c> fragments. Do not parse and
    # reserialize the full worksheet because Excel workbooks often contain
    # extension namespaces that must remain byte-for-byte intact.
    text = xml_bytes.decode("utf-8")
    for change in changes:
        cell_ref = re.escape(str(change["cell"]))
        pattern = re.compile(rf'(?s)<c\b(?=[^>]*\br="{cell_ref}")[^>]*/>|<c\b(?=[^>]*\br="{cell_ref}")[^>]*>.*?</c>')
        text = pattern.sub(lambda m, value=change["new_value"]: _patch_cell_fragment(m.group(0), value), text, count=1)
    return text.encode("utf-8")


def _remove_calc_chain_content_type(xml_bytes: bytes) -> bytes:
    text = xml_bytes.decode("utf-8")
    text = re.sub(r'<Override\b(?=[^>]*PartName="/xl/calcChain\.xml")[^>]*/>', "", text)
    return text.encode("utf-8")


def _remove_calc_chain_relationship(xml_bytes: bytes) -> bytes:
    text = xml_bytes.decode("utf-8")
    text = re.sub(r'<Relationship\b(?=[^>]*Type="[^"]*/calcChain")[^>]*/>', "", text)
    return text.encode("utf-8")


def _set_tag_attr(tag: str, name: str, value: str) -> str:
    tag = re.sub(rf'\s{name}="[^"]*"', "", tag)
    if tag.endswith("/>"):
        return tag[:-2] + f' {name}="{value}"/>'
    return tag[:-1] + f' {name}="{value}">'


def _force_full_recalc_workbook_xml(xml_bytes: bytes) -> bytes:
    text = xml_bytes.decode("utf-8")
    match = re.search(r"<calcPr\b[^>]*/>|<calcPr\b[^>]*>.*?</calcPr>", text, flags=re.S)
    if match:
        tag_match = re.match(r"<calcPr\b[^>]*>", match.group(0))
        if not tag_match:
            return xml_bytes
        tag = tag_match.group(0)
        for name, value in [
            ("calcMode", "auto"),
            ("fullCalcOnLoad", "1"),
            ("forceFullCalc", "1"),
            ("calcOnSave", "1"),
        ]:
            tag = _set_tag_attr(tag, name, value)
        replacement = tag if match.group(0).endswith("/>") else tag + match.group(0)[tag_match.end() :]
        text = text[: match.start()] + replacement + text[match.end() :]
    else:
        insert_at = text.rfind("</workbook>")
        if insert_at != -1:
            text = (
                text[:insert_at]
                + '<calcPr calcMode="auto" fullCalcOnLoad="1" forceFullCalc="1" calcOnSave="1"/>'
                + text[insert_at:]
            )
    return text.encode("utf-8")


def _write_patched_workbook(original_path: Path, out_path: Path, changes: list[dict[str, Any]]) -> None:
    changes_by_sheet: dict[str, list[dict[str, Any]]] = {}
    for change in changes:
        changes_by_sheet.setdefault(change["sheet"], []).append(change)

    sheet_paths = _sheet_xml_paths_by_name(original_path)
    patch_by_part = {
        sheet_paths[sheet]: sheet_changes
        for sheet, sheet_changes in changes_by_sheet.items()
        if sheet in sheet_paths
    }

    temp_path = out_path.with_suffix(out_path.suffix + ".tmp")
    with zipfile.ZipFile(original_path, "r") as zin, zipfile.ZipFile(temp_path, "w") as zout:
        for item in zin.infolist():
            if item.filename == "xl/calcChain.xml":
                continue
            data = zin.read(item.filename)
            if item.filename in patch_by_part:
                data = _patch_sheet_xml(data, patch_by_part[item.filename])
            elif item.filename == "[Content_Types].xml":
                data = _remove_calc_chain_content_type(data)
            elif item.filename == "xl/_rels/workbook.xml.rels":
                data = _remove_calc_chain_relationship(data)
            elif item.filename == "xl/workbook.xml":
                data = _force_full_recalc_workbook_xml(data)
            zout.writestr(item, data)
    temp_path.replace(out_path)


def export_workbook_value_transfer(
    original_path: Path,
    edited_path: Path,
    output_dir: Path,
    sheet_name: str | None = None,
    include_formula_cells: bool = False,
    include_blank_cells: bool = False,
    output_name: str | None = None,
) -> tuple[Path, dict[str, Any]]:
    plan = plan_workbook_value_transfer(
        original_path=original_path,
        edited_path=edited_path,
        sheet_name=sheet_name,
        include_formula_cells=include_formula_cells,
        include_blank_cells=include_blank_cells,
        max_preview=10**9,
    )

    output_dir.mkdir(parents=True, exist_ok=True)
    stem = safe_filename(output_name or f"{original_path.stem}_updated_from_{edited_path.stem}")
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = output_dir / f"{stamp}_{stem}{original_path.suffix}"

    if plan["change_count"] == 0:
        shutil.copy2(original_path, out_path)
        plan["output_file_note"] = (
            "No cells were patched. The exported file is an untouched copy of the original workbook."
        )
        plan["changes_preview"] = []
        plan["preview_limit"] = 200
        return out_path, plan

    all_changes = plan["changes_preview"]
    _write_patched_workbook(original_path, out_path, all_changes)
    plan["changes_preview"] = all_changes[:200]
    plan["preview_limit"] = 200
    plan["preview_truncated"] = len(all_changes) > 200

    return out_path, plan
