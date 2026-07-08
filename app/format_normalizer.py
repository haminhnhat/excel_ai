from __future__ import annotations

import shutil
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .utils import normalize_text, safe_filename


SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm"}


def _validate_workbook_path(path: Path, label: str) -> None:
    if not path.exists():
        raise FileNotFoundError(f"{label} workbook not found: {path}")
    if path.suffix.lower() not in SUPPORTED_EXTENSIONS:
        raise ValueError(f"{label} workbook must be .xlsx or .xlsm: {path.name}")


def _workbook_kwargs(path: Path) -> dict[str, Any]:
    return {"keep_vba": path.suffix.lower() == ".xlsm"}


def _norm(value: Any) -> str:
    return normalize_text(str(value or "")).replace("_", " ").strip()


def _header_score(values: list[Any]) -> int:
    score = 0
    for value in values:
        if isinstance(value, str) and value.strip():
            score += 3
        elif value not in (None, ""):
            score += 1
    return score


def detect_header_row(ws: Any, max_scan_rows: int = 40, min_headers: int = 2) -> tuple[int, dict[int, str]]:
    best_row = 0
    best_score = 0
    best_headers: dict[int, str] = {}
    max_row = min(ws.max_row, max_scan_rows)
    for row in range(1, max_row + 1):
        values = [ws.cell(row=row, column=col).value for col in range(1, ws.max_column + 1)]
        headers = {
            col: str(value).strip()
            for col, value in enumerate(values, start=1)
            if isinstance(value, str) and str(value).strip()
        }
        if len(headers) < min_headers:
            continue
        score = _header_score(values)
        if score > best_score:
            best_row = row
            best_score = score
            best_headers = headers
    if not best_headers:
        raise ValueError(f"Could not detect a table header row in sheet: {ws.title}")
    return best_row, best_headers


def _match_score(target: str, candidate: str) -> float:
    target_norm = _norm(target)
    candidate_norm = _norm(candidate)
    if not target_norm or not candidate_norm:
        return 0.0
    if target_norm == candidate_norm:
        return 1.0
    if target_norm in candidate_norm or candidate_norm in target_norm:
        return 0.9
    target_tokens = set(target_norm.split())
    candidate_tokens = set(candidate_norm.split())
    token_score = 0.0
    if target_tokens and candidate_tokens:
        token_score = len(target_tokens & candidate_tokens) / len(target_tokens | candidate_tokens)
    fuzzy = SequenceMatcher(None, target_norm, candidate_norm).ratio()
    return max(fuzzy, token_score)


def _best_match(target: str, candidates: dict[int, str]) -> tuple[int | None, str | None, float]:
    best_col: int | None = None
    best_header: str | None = None
    best_score = 0.0
    for col, header in candidates.items():
        score = _match_score(target, header)
        if score > best_score:
            best_col = col
            best_header = header
            best_score = score
    return best_col, best_header, best_score


def _match_sheet(template_sheet: str, source_sheets: list[str]) -> tuple[str | None, float]:
    best_sheet: str | None = None
    best_score = 0.0
    for sheet in source_sheets:
        score = _match_score(template_sheet, sheet)
        if score > best_score:
            best_sheet = sheet
            best_score = score
    return best_sheet, best_score


def _confidence(score: float) -> str:
    if score >= 0.9:
        return "high"
    if score >= 0.72:
        return "medium"
    return "low"


def _data_rows(ws: Any, header_row: int, header_cols: list[int], max_blank_streak: int = 20) -> list[int]:
    rows: list[int] = []
    blank_streak = 0
    for row in range(header_row + 1, ws.max_row + 1):
        has_value = any(ws.cell(row=row, column=col).value not in (None, "") for col in header_cols)
        if has_value:
            rows.append(row)
            blank_streak = 0
        else:
            blank_streak += 1
            if blank_streak >= max_blank_streak:
                break
    return rows


def preview_format_normalization(
    source_path: Path,
    template_path: Path,
    sheet_name: str | None = None,
    min_confidence: float = 0.72,
) -> dict[str, Any]:
    _validate_workbook_path(source_path, "Source")
    _validate_workbook_path(template_path, "Template")

    source_wb = load_workbook(source_path, data_only=True, read_only=False, **_workbook_kwargs(source_path))
    template_wb = load_workbook(template_path, data_only=False, read_only=False, **_workbook_kwargs(template_path))
    try:
        template_sheets = [sheet_name] if sheet_name else template_wb.sheetnames
        sheet_previews: list[dict[str, Any]] = []
        for template_sheet in template_sheets:
            if template_sheet not in template_wb.sheetnames:
                raise ValueError(f"Sheet not found in template workbook: {template_sheet}")
            source_sheet, sheet_score = _match_sheet(template_sheet, source_wb.sheetnames)
            if source_sheet is None or sheet_score < 0.45:
                sheet_previews.append({
                    "template_sheet": template_sheet,
                    "source_sheet": None,
                    "sheet_confidence": "low",
                    "mappings": [],
                    "warnings": ["No likely matching source sheet found."],
                })
                continue

            template_ws = template_wb[template_sheet]
            source_ws = source_wb[source_sheet]
            template_header_row, template_headers = detect_header_row(template_ws)
            source_header_row, source_headers = detect_header_row(source_ws)
            source_data_rows = _data_rows(source_ws, source_header_row, list(source_headers.keys()))

            mappings = []
            for target_col, target_header in template_headers.items():
                source_col, source_header, score = _best_match(target_header, source_headers)
                mappings.append({
                    "template_header": target_header,
                    "template_col": target_col,
                    "source_header": source_header,
                    "source_col": source_col,
                    "score": round(score, 3),
                    "confidence": _confidence(score),
                    "will_copy": bool(source_col is not None and score >= min_confidence),
                })

            sheet_previews.append({
                "template_sheet": template_sheet,
                "source_sheet": source_sheet,
                "sheet_score": round(sheet_score, 3),
                "sheet_confidence": _confidence(sheet_score),
                "template_header_row": template_header_row,
                "source_header_row": source_header_row,
                "source_data_row_count": len(source_data_rows),
                "mappings": mappings,
                "warnings": [
                    "Some columns need review before export."
                ] if any(not m["will_copy"] for m in mappings) else [],
            })

        return {
            "ok": True,
            "source_file": source_path.name,
            "template_file": template_path.name,
            "min_confidence": min_confidence,
            "sheets": sheet_previews,
        }
    finally:
        source_wb.close()
        template_wb.close()


def export_format_normalization(
    source_path: Path,
    template_path: Path,
    output_dir: Path,
    sheet_name: str | None = None,
    min_confidence: float = 0.72,
    output_name: str | None = None,
) -> tuple[Path, dict[str, Any]]:
    preview = preview_format_normalization(
        source_path=source_path,
        template_path=template_path,
        sheet_name=sheet_name,
        min_confidence=min_confidence,
    )
    output_dir.mkdir(parents=True, exist_ok=True)
    stem = safe_filename(output_name or f"{source_path.stem}_normalized_to_{template_path.stem}")
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = output_dir / f"{stamp}_{stem}{template_path.suffix}"
    shutil.copy2(template_path, out_path)

    source_wb = load_workbook(source_path, data_only=True, read_only=False, **_workbook_kwargs(source_path))
    out_wb = load_workbook(out_path, data_only=False, read_only=False, **_workbook_kwargs(out_path))
    try:
        for sheet_preview in preview["sheets"]:
            if not sheet_preview.get("source_sheet") or sheet_preview.get("template_sheet") not in out_wb.sheetnames:
                continue
            source_ws = source_wb[sheet_preview["source_sheet"]]
            out_ws = out_wb[sheet_preview["template_sheet"]]
            source_header_row = sheet_preview["source_header_row"]
            template_header_row = sheet_preview["template_header_row"]
            rows = _data_rows(
                source_ws,
                source_header_row,
                [
                    item["source_col"]
                    for item in sheet_preview["mappings"]
                    if item.get("source_col")
                ],
            )

            for out_offset, source_row in enumerate(rows, start=1):
                out_row = template_header_row + out_offset
                for mapping in sheet_preview["mappings"]:
                    if not mapping["will_copy"]:
                        continue
                    out_ws.cell(row=out_row, column=mapping["template_col"]).value = source_ws.cell(
                        row=source_row,
                        column=mapping["source_col"],
                    ).value

        out_wb.save(out_path)
        return out_path, preview
    finally:
        source_wb.close()
        out_wb.close()
