from __future__ import annotations

import os
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Dict, Tuple

from .schemas import ActionPlan, AppliedChange, OutputValue, ScenarioResult
from .utils import append_audit_log, safe_filename


class ExcelControllerError(Exception):
    pass


class ExcelController:
    def __init__(
        self,
        model_map: Dict[str, Any],
        output_dir: str | Path = "outputs",
        audit_log_path: str | Path = "logs/audit_log.csv",
        engine: str = "auto",
        progress_callback: Callable[[str], None] | None = None,
        save_workbook: bool = True,
        write_audit_log: bool = True,
    ) -> None:
        self.model_map = model_map
        self.output_dir = Path(output_dir)
        self.audit_log_path = Path(audit_log_path)
        self.engine = engine
        self.progress_callback = progress_callback
        self.save_workbook = save_workbook
        self.write_audit_log = write_audit_log
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.audit_log_path.parent.mkdir(parents=True, exist_ok=True)

    def _progress(self, stage: str) -> None:
        if self.progress_callback is not None:
            self.progress_callback(stage)

    def run_scenario(self, source_file: str | Path, plan: ActionPlan) -> ScenarioResult:
        source = Path(source_file)
        if not source.exists():
            raise FileNotFoundError(f"Excel source file not found: {source}")

        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        scenario_name = safe_filename(plan.scenario_name)
        target = self.output_dir / f"{stamp}_{scenario_name}{source.suffix}"
        self._progress("copy_workbook:start")
        shutil.copy2(source, target)
        self._progress("copy_workbook:done")

        engine = self.engine.lower().strip()
        warnings: list[str] = []

        if engine == "auto":
            try:
                self._progress("engine:xlwings:start")
                applied, outputs = self._run_with_xlwings(target, plan)
                used_engine = "xlwings"
                recalculated = True
            except Exception as e:
                self._progress(f"engine:xlwings:failed:{e}")
                warnings.append(f"xlwings failed, falling back to openpyxl. Reason: {e}")
                self._progress("engine:openpyxl:start")
                applied, outputs = self._run_with_openpyxl(target, plan)
                used_engine = "openpyxl"
                recalculated = False
        elif engine == "xlwings":
            self._progress("engine:xlwings:start")
            applied, outputs = self._run_with_xlwings(target, plan)
            used_engine = "xlwings"
            recalculated = True
        elif engine == "openpyxl":
            self._progress("engine:openpyxl:start")
            applied, outputs = self._run_with_openpyxl(target, plan)
            used_engine = "openpyxl"
            recalculated = False
            warnings.append(
                "openpyxl edited the workbook but did not calculate Excel formulas. "
                "Open the scenario file in Excel to force recalculation, or use engine=xlwings."
            )
        else:
            raise ExcelControllerError(f"Unsupported engine: {self.engine}")

        if self.write_audit_log:
            append_audit_log(
                self.audit_log_path,
                [
                    {
                        "scenario_name": plan.scenario_name,
                        "parameter": c.parameter,
                        "sheet": c.sheet,
                        "cell": c.cell,
                        "old_value": c.old_value,
                        "new_value": c.new_value,
                        "reason": c.reason,
                    }
                    for c in applied
                ],
            )
            self._progress("audit_log:done")
        else:
            self._progress("audit_log:skipped:preview")

        return ScenarioResult(
            scenario_name=plan.scenario_name,
            source_file=str(source),
            scenario_file=str(target),
            engine=used_engine,
            recalculated=recalculated,
            applied_changes=applied,
            outputs=outputs,
            warnings=warnings,
        )

    def _target_value(self, old_value: Any, operation: str, value: Any) -> Any:
        if operation == "set":
            return value
        if not isinstance(old_value, (int, float)) or not isinstance(value, (int, float)):
            raise ExcelControllerError(f"Operation {operation} requires numeric old and new values.")
        if operation == "add":
            return old_value + value
        if operation == "multiply":
            return old_value * value
        raise ExcelControllerError(f"Unsupported operation: {operation}")

    def _check_formula_overwrite(self, formula: Any, sheet: str, cell: str, allow_cell: bool = False) -> None:
        allow = allow_cell or self.model_map.get("settings", {}).get("allow_formula_overwrite", False)
        if allow:
            return
        if isinstance(formula, str) and formula.startswith("="):
            raise ExcelControllerError(f"Refusing to overwrite formula cell: {sheet}!{cell}")

    def _input_targets(self, meta: Dict[str, Any]) -> list[str]:
        """Return target cell addresses for one input mapping.

        A mapping can use either:
        - cell: "D12"
        - cells: ["D12", "E12"]
        - range: "D12:H12"

        The same new value is applied to all target cells. This is useful for
        financial models where one assumption, such as loan interest rate, is
        hardcoded across multiple period columns.
        """
        if meta.get("cell"):
            return [str(meta["cell"])]
        if meta.get("cells"):
            return [str(c) for c in meta["cells"]]
        if meta.get("range"):
            from openpyxl.utils import get_column_letter, range_boundaries

            min_col, min_row, max_col, max_row = range_boundaries(str(meta["range"]))
            cells: list[str] = []
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    cells.append(f"{get_column_letter(col)}{row}")
            return cells
        raise ExcelControllerError("Input mapping must define one of: cell, cells, range.")

    def _run_with_xlwings(self, workbook_path: Path, plan: ActionPlan) -> Tuple[list[AppliedChange], Dict[str, OutputValue]]:
        import xlwings as xw

        self._progress("xlwings:create_excel_app:start")
        app = xw.App(visible=False, add_book=False)
        app.display_alerts = False
        app.screen_updating = False
        book = None
        try:
            self._progress("xlwings:open_workbook:start")
            book = app.books.open(str(workbook_path))
            self._progress("xlwings:open_workbook:done")
            applied: list[AppliedChange] = []

            self._progress("xlwings:write_inputs:start")
            for change in plan.changes:
                meta = self.model_map["inputs"][change.parameter]
                sheet_name = meta["sheet"]
                for cell_addr in self._input_targets(meta):
                    rng = book.sheets[sheet_name].range(cell_addr)
                    old_value = rng.value
                    self._check_formula_overwrite(rng.formula, sheet_name, cell_addr, meta.get("allow_formula_overwrite", False))
                    new_value = self._target_value(old_value, change.operation, change.value)
                    rng.value = new_value
                    applied.append(
                        AppliedChange(
                            parameter=change.parameter,
                            sheet=sheet_name,
                            cell=cell_addr,
                            old_value=old_value,
                            new_value=new_value,
                            reason=change.reason,
                        )
                    )
            self._progress(f"xlwings:write_inputs:done:{len(applied)}")

            # Force full recalc only when it is useful. Read-only requests can use
            # workbook values directly; recalculating large Excel models can take minutes.
            recalc_mode = os.getenv("EXCEL_RECALC_MODE", "calculate").lower().strip()
            if not applied and os.getenv("EXCEL_RECALC_READ_ONLY", "false").lower().strip() != "true":
                recalc_mode = "none"
            self._progress(f"xlwings:recalculate:start:{recalc_mode}")
            if recalc_mode == "full_rebuild":
                try:
                    app.api.CalculateFullRebuild()
                except Exception:
                    app.calculate()
            elif recalc_mode == "full":
                try:
                    app.api.CalculateFull()
                except Exception:
                    app.calculate()
            elif recalc_mode == "none":
                self._progress("xlwings:recalculate:skipped")
            else:
                app.calculate()
            self._progress(f"xlwings:recalculate:done:{recalc_mode}")

            self._progress("xlwings:read_outputs:start")
            outputs = self._read_outputs_xlwings(book, plan)
            self._progress(f"xlwings:read_outputs:done:{len(outputs)}")
            if applied and self.save_workbook:
                self._progress("xlwings:save_workbook:start")
                book.save(str(workbook_path))
                self._progress("xlwings:save_workbook:done")
            elif applied:
                self._progress("xlwings:save_workbook:skipped:preview")
            else:
                self._progress("xlwings:save_workbook:skipped:no_changes")
            return applied, outputs
        finally:
            if book is not None:
                self._progress("xlwings:close_workbook")
                book.close()
            self._progress("xlwings:quit_excel")
            app.quit()

    def _read_outputs_xlwings(self, book: Any, plan: ActionPlan) -> Dict[str, OutputValue]:
        outputs: Dict[str, OutputValue] = {}
        output_map = self.model_map.get("outputs", {})
        for key in plan.requested_outputs:
            meta = output_map[key]
            value = book.sheets[meta["sheet"]].range(meta["cell"]).value
            outputs[key] = OutputValue(
                key=key,
                sheet=meta["sheet"],
                cell=meta["cell"],
                value=value,
                type=meta.get("type"),
                unit=meta.get("unit"),
                description=meta.get("description"),
            )
        return outputs

    def _run_with_openpyxl(self, workbook_path: Path, plan: ActionPlan) -> Tuple[list[AppliedChange], Dict[str, OutputValue]]:
        from openpyxl import load_workbook

        self._progress("openpyxl:load_workbook:start")
        wb = load_workbook(workbook_path)
        self._progress("openpyxl:load_workbook:done")
        applied: list[AppliedChange] = []

        self._progress("openpyxl:write_inputs:start")
        for change in plan.changes:
            meta = self.model_map["inputs"][change.parameter]
            sheet_name = meta["sheet"]
            ws = wb[sheet_name]
            for cell_addr in self._input_targets(meta):
                cell = ws[cell_addr]
                old_value = cell.value
                formula = old_value if isinstance(old_value, str) and old_value.startswith("=") else None
                self._check_formula_overwrite(formula, sheet_name, cell_addr, meta.get("allow_formula_overwrite", False))
                new_value = self._target_value(old_value, change.operation, change.value)
                cell.value = new_value
                applied.append(
                    AppliedChange(
                        parameter=change.parameter,
                        sheet=sheet_name,
                        cell=cell_addr,
                        old_value=old_value,
                        new_value=new_value,
                        reason=change.reason,
                    )
                )
        self._progress(f"openpyxl:write_inputs:done:{len(applied)}")

        # Ask Excel to recalculate when the workbook is opened.
        try:
            wb.calculation.fullCalcOnLoad = True
            wb.calculation.forceFullCalc = True
            wb.calculation.calcMode = "auto"
        except Exception:
            pass
        self._progress("openpyxl:save_workbook:start")
        wb.save(workbook_path)
        self._progress("openpyxl:save_workbook:done")
        wb.close()

        # Read cached formula results. These may be stale unless Excel recalculated previously.
        self._progress("openpyxl:read_cached_outputs:start")
        wb_values = load_workbook(workbook_path, data_only=True)
        outputs: Dict[str, OutputValue] = {}
        for key in plan.requested_outputs:
            meta = self.model_map["outputs"][key]
            value = wb_values[meta["sheet"]][meta["cell"]].value
            outputs[key] = OutputValue(
                key=key,
                sheet=meta["sheet"],
                cell=meta["cell"],
                value=value,
                type=meta.get("type"),
                unit=meta.get("unit"),
                description=meta.get("description"),
            )
        wb_values.close()
        self._progress(f"openpyxl:read_cached_outputs:done:{len(outputs)}")
        return applied, outputs
