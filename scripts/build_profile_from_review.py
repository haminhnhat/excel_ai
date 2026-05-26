from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

import yaml
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from app.utils import safe_filename


def split_aliases(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, list):
        return [str(v).strip() for v in value if str(v).strip()]
    text = str(value)
    parts = []
    for item in text.replace(";", "|").replace(",", "|").split("|"):
        item = item.strip()
        if item:
            parts.append(item)
    return sorted(set(parts), key=lambda x: x.lower())


def clean_bool(value: Any, default: bool = True) -> bool:
    if value is None:
        return default
    text = str(value).strip().lower()
    if text in {"yes", "y", "true", "1", "approve", "approved"}:
        return True
    if text in {"no", "n", "false", "0", "reject", "rejected"}:
        return False
    return default


def numeric_or_none(value: Any) -> Any:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return value
    try:
        return float(str(value).strip())
    except ValueError:
        return None


def normalize_decision(value: Any) -> str:
    return str(value or "").strip().lower()


def find_col(headers: list[str], name: str) -> int:
    normalized = [str(h).strip().lower() for h in headers]
    return normalized.index(name.lower())


def read_review_rows(review_path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(review_path, data_only=True)
    if "Review Mapping" not in wb.sheetnames:
        raise ValueError("Review workbook must contain a 'Review Mapping' sheet.")
    ws = wb["Review Mapping"]
    headers = [c.value for c in ws[1]]
    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        data = {str(headers[i]).strip(): row[i] for i in range(min(len(headers), len(row))) if headers[i]}
        rows.append(data)
    return rows


def target_fields(row: dict[str, Any]) -> dict[str, Any]:
    sheet = str(row.get("Correct Sheet") or row.get("Sheet") or "").strip()
    target = str(row.get("Correct Cell/Range") or row.get("Cell") or "").strip()
    mode = str(row.get("Target Mode") or "cell").strip().lower()
    if not sheet or not target:
        raise ValueError(f"Approved row lacks sheet/cell: {row}")
    fields: dict[str, Any] = {"sheet": sheet}
    if mode == "range" or ":" in target:
        fields["range"] = target
    elif mode == "cells" or "," in target:
        fields["cells"] = [c.strip() for c in target.split(",") if c.strip()]
    else:
        fields["cell"] = target
    return fields


def build_map_from_review(review_path: Path, profile_name: str, currency: str = "VND") -> dict[str, Any]:
    rows = read_review_rows(review_path)
    inputs: dict[str, Any] = {}
    outputs: dict[str, Any] = {}
    rejected = 0
    skipped = 0

    for row in rows:
        decision = normalize_decision(row.get("Decision"))
        if decision in {"reject", "rejected", "no"}:
            rejected += 1
            continue
        if decision not in {"approve", "approved", "yes"}:
            skipped += 1
            continue
        role = str(row.get("Role") or "").strip().lower()
        key = str(row.get("Parameter Key") or "").strip()
        if not key:
            skipped += 1
            continue
        fields = target_fields(row)
        type_ = str(row.get("Type") or "number").strip() or "number"
        unit = str(row.get("Unit") or "").strip() or None
        description = str(row.get("Description") or row.get("Friendly Name") or key).strip()
        aliases = split_aliases(row.get("Aliases"))

        if role == "input":
            meta = {
                **fields,
                "type": type_,
                "editable": clean_bool(row.get("Editable"), True),
                "min": numeric_or_none(row.get("Min")),
                "max": numeric_or_none(row.get("Max")),
                "unit": unit,
                "description": description,
                "aliases": aliases,
            }
            # remove null fields so YAML stays clean
            inputs[key] = {k: v for k, v in meta.items() if v is not None and v != []}
        elif role == "output":
            if "range" in fields or "cells" in fields:
                raise ValueError(f"Output {key} must map to a single cell, not range/cells: {fields}")
            meta = {
                **fields,
                "type": type_,
                "unit": unit,
                "description": description,
                "aliases": aliases,
            }
            outputs[key] = {k: v for k, v in meta.items() if v is not None and v != []}
        else:
            skipped += 1

    default_order = [
        "profit_after_tax", "project_npv", "project_irr", "equity_npv", "equity_irr", "roi", "total_investment", "bank_loan", "total_revenue", "revenue"
    ]
    default_outputs = [k for k in default_order if k in outputs] or list(outputs.keys())[:8]
    model_map = {
        "metadata": {
            "name": profile_name,
            "version": "review-approved-0.1",
            "currency": currency,
            "source_review": str(review_path),
            "note": "Generated from mapping_review.xlsx. Validate before production use.",
        },
        "settings": {
            "allow_formula_overwrite": False,
            "default_outputs": default_outputs,
        },
        "inputs": inputs,
        "outputs": outputs,
    }
    model_map["metadata"]["review_stats"] = {
        "approved_inputs": len(inputs),
        "approved_outputs": len(outputs),
        "rejected_rows": rejected,
        "skipped_review_rows": skipped,
    }
    return model_map


def main() -> None:
    parser = argparse.ArgumentParser(description="Build model_map.yaml from an approved mapping_review.xlsx workbook.")
    parser.add_argument("--review", required=True, help="Path to mapping_review.xlsx")
    parser.add_argument("--profile", required=True, help="Profile name to create under config/profiles/<profile>")
    parser.add_argument("--currency", default="VND")
    parser.add_argument("--force", action="store_true", help="Overwrite existing profile model_map.yaml")
    args = parser.parse_args()

    review_path = Path(args.review)
    if not review_path.is_absolute():
        review_path = ROOT / review_path
    if not review_path.exists():
        raise FileNotFoundError(review_path)

    profile = safe_filename(args.profile)
    profile_dir = ROOT / "config" / "profiles" / profile
    out_path = profile_dir / "model_map.yaml"
    if out_path.exists() and not args.force:
        raise FileExistsError(f"Profile already exists: {out_path}. Use --force to overwrite.")
    profile_dir.mkdir(parents=True, exist_ok=True)

    model_map = build_map_from_review(review_path, profile_name=profile, currency=args.currency)
    out_path.write_text(yaml.safe_dump(model_map, allow_unicode=True, sort_keys=False), encoding="utf-8")
    readme = profile_dir / "README.md"
    readme.write_text(
        f"# Profile: {profile}\n\n"
        f"Generated from `{review_path}`.\n\n"
        "Run validation before using it:\n\n"
        f"```powershell\npython scripts/validate_profile.py --profile {profile} --excel \"models/your_model.xlsx\"\n```\n",
        encoding="utf-8",
    )
    summary = {
        "profile": profile,
        "model_map": str(out_path),
        "inputs": len(model_map.get("inputs", {})),
        "outputs": len(model_map.get("outputs", {})),
        "next_step": f"python scripts/validate_profile.py --profile {profile} --excel \"models/your_model.xlsx\"",
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
