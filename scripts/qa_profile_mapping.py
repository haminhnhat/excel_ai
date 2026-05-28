from __future__ import annotations

import argparse
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from app.config_loader import load_model_map_for_profile


@dataclass
class Finding:
    severity: str
    key: str
    message: str


def _targets(meta: dict[str, Any]) -> list[str]:
    if meta.get("cell"):
        return [str(meta["cell"])]
    if meta.get("cells"):
        return [str(cell) for cell in meta["cells"]]
    if meta.get("range"):
        return [str(meta["range"])]
    return []


def _check_static(model_map: dict[str, Any]) -> list[Finding]:
    findings: list[Finding] = []
    inputs = model_map.get("inputs", {})
    outputs = model_map.get("outputs", {})
    defaults = model_map.get("settings", {}).get("default_outputs", [])

    for key, meta in inputs.items():
        if not meta.get("sheet"):
            findings.append(Finding("ERROR", f"inputs.{key}", "Thiếu sheet."))
        if not _targets(meta):
            findings.append(Finding("ERROR", f"inputs.{key}", "Thiếu cell/cells/range."))
        if meta.get("editable", True) and (meta.get("min") is None or meta.get("max") is None):
            findings.append(Finding("WARN", f"inputs.{key}", "Input editable nên có min và max."))
        if not meta.get("aliases"):
            findings.append(Finding("WARN", f"inputs.{key}", "Thiếu aliases tiếng Việt/ngôn ngữ nội bộ."))
        if not meta.get("description"):
            findings.append(Finding("INFO", f"inputs.{key}", "Nên có description để review dễ hơn."))

    for key, meta in outputs.items():
        if not meta.get("sheet"):
            findings.append(Finding("ERROR", f"outputs.{key}", "Thiếu sheet."))
        if not meta.get("cell"):
            findings.append(Finding("ERROR", f"outputs.{key}", "Output nên có cell đơn."))
        if not meta.get("aliases"):
            findings.append(Finding("WARN", f"outputs.{key}", "Thiếu aliases cho câu hỏi output."))
        if not meta.get("description"):
            findings.append(Finding("INFO", f"outputs.{key}", "Nên có description để người dùng hiểu output."))

    for key in defaults:
        if key not in outputs:
            findings.append(Finding("ERROR", f"settings.default_outputs.{key}", "Default output không tồn tại trong outputs."))

    return findings


def _check_workbook(model_map: dict[str, Any], excel_path: Path) -> list[Finding]:
    from openpyxl import load_workbook

    findings: list[Finding] = []
    wb = load_workbook(excel_path, data_only=False, read_only=False)
    try:
        for section in ("inputs", "outputs"):
            for key, meta in model_map.get(section, {}).items():
                sheet_name = meta.get("sheet")
                if sheet_name not in wb.sheetnames:
                    findings.append(Finding("ERROR", f"{section}.{key}", f"Sheet không tồn tại: {sheet_name}"))
                    continue
                ws = wb[sheet_name]
                for target in _targets(meta):
                    if ":" in target:
                        continue
                    try:
                        value = ws[target].value
                    except Exception as e:
                        findings.append(Finding("ERROR", f"{section}.{key}", f"Cell không đọc được {sheet_name}!{target}: {e}"))
                        continue
                    is_formula = isinstance(value, str) and value.startswith("=")
                    if section == "inputs" and is_formula and not meta.get("allow_formula_overwrite", False):
                        findings.append(Finding("ERROR", f"inputs.{key}", f"Input đang trỏ vào ô công thức: {sheet_name}!{target}"))
                    if section == "outputs" and value is None:
                        findings.append(Finding("WARN", f"outputs.{key}", f"Output đang rỗng trong workbook: {sheet_name}!{target}"))
    finally:
        wb.close()
    return findings


def main() -> int:
    parser = argparse.ArgumentParser(description="QA model_map profile for output/input accuracy risks.")
    parser.add_argument("--profile", required=True)
    parser.add_argument("--excel", default=None, help="Optional workbook path to verify sheets/cells/formulas.")
    args = parser.parse_args()

    model_map, model_map_path = load_model_map_for_profile(profile=args.profile)
    findings = _check_static(model_map)
    if args.excel:
        findings.extend(_check_workbook(model_map, Path(args.excel)))

    print(f"Profile QA: {args.profile}")
    print(f"Model map: {model_map_path}")
    if args.excel:
        print(f"Workbook: {args.excel}")
    print()

    if not findings:
        print("PASS Không phát hiện vấn đề mapping rõ ràng.")
        return 0

    counts = {"ERROR": 0, "WARN": 0, "INFO": 0}
    for finding in findings:
        counts[finding.severity] = counts.get(finding.severity, 0) + 1
        print(f"{finding.severity} {finding.key}: {finding.message}")

    print()
    print(f"Summary: {counts.get('ERROR', 0)} error, {counts.get('WARN', 0)} warn, {counts.get('INFO', 0)} info")
    return 1 if counts.get("ERROR", 0) else 0


if __name__ == "__main__":
    raise SystemExit(main())
