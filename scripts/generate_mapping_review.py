from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from app.utils import safe_filename
from analyze_workbook import (  # type: ignore
    INPUT_KEYWORDS,
    OUTPUT_KEYWORDS,
    ERROR_TOKENS,
    cell_addr,
    compact_value,
    contains_error,
    is_formula,
    nearby_label,
    text_matches,
    value_type_guess,
)

FRIENDLY_NAMES = {
    "loan_interest_rate": "Lãi vay ngân hàng",
    "vat_rate": "Thuế VAT",
    "cit_rate": "Thuế TNDN / CIT",
    "selling_price_change": "Điều chỉnh giá bán",
    "investment_cost_change": "Điều chỉnh tổng mức đầu tư / TMĐT",
    "selling_cost_rate": "Chi phí bán hàng / môi giới",
    "marketing_rate": "Chi phí marketing",
    "admin_cost_rate": "Chi phí quản lý",
    "sales_volume_change": "Điều chỉnh sản lượng bán",
    "debt_financing_ratio": "Tỷ lệ vay / debt ratio",
    "project_npv": "NPV dự án",
    "project_irr": "IRR dự án",
    "equity_npv": "NPV chủ đầu tư",
    "equity_irr": "IRR chủ đầu tư",
    "roi": "ROI",
    "profit_after_tax": "Lợi nhuận sau thuế",
    "total_investment": "Tổng mức đầu tư",
    "revenue": "Doanh thu",
    "total_revenue": "Tổng doanh thu",
    "bank_loan": "Vay ngân hàng",
}

REVIEW_HEADERS = [
    "Decision",
    "Role",
    "Parameter Key",
    "Friendly Name",
    "Sheet",
    "Cell",
    "Current Value",
    "Formula?",
    "Type",
    "Unit",
    "Min",
    "Max",
    "Editable",
    "Confidence",
    "Nearby Label",
    "Description",
    "Aliases",
    "Correct Sheet",
    "Correct Cell/Range",
    "Target Mode",
    "Notes",
]

INPUT_HEADERS = [
    "Parameter Guess", "Sheet", "Cell", "Value", "Type Guess", "Formula?", "Confidence", "Nearby Label"
]
OUTPUT_HEADERS = [
    "Output Guess", "Sheet", "Cell", "Value/Formula", "Type Guess", "Formula?", "Confidence", "Nearby Label"
]
ERROR_HEADERS = ["Sheet", "Cell", "Value/Formula", "Nearby Label"]

HEADER_FILL = "1F4E79"
HEADER_FONT = "FFFFFF"
LIGHT_FILL = "D9EAF7"
INPUT_FILL = "E2F0D9"
OUTPUT_FILL = "FCE4D6"
WARN_FILL = "FFF2CC"
ERROR_FILL = "F4CCCC"
BORDER = Border(bottom=Side(style="thin", color="D9E2F3"))


def guess_confidence(param: str, label: str, role: str, is_formula_cell: bool, value: Any) -> str:
    if not param:
        return "low"
    label_text = str(label or "")
    phrases = INPUT_KEYWORDS.get(param, []) if role == "input" else OUTPUT_KEYWORDS.get(param, [])
    matched = any(text_matches(label_text, [p]) for p in phrases)
    if role == "input":
        if matched and not is_formula_cell:
            return "high"
        if not is_formula_cell:
            return "medium"
        return "low"
    if role == "output":
        if matched and is_formula_cell:
            return "high"
        if matched:
            return "medium"
        return "low"
    return "low"


def default_min_max(param: str, type_guess: str) -> tuple[Any, Any]:
    if param == "loan_interest_rate":
        return 0.0, 0.25
    if param == "vat_rate":
        return 0.0, 0.20
    if param == "cit_rate":
        return 0.0, 0.35
    if param in {"selling_price_change", "sales_volume_change"}:
        return -0.50, 1.00
    if param == "investment_cost_change":
        return -0.30, 1.00
    if type_guess == "percent":
        return -1.0, 1.0
    return None, None


def aliases_for(param: str, role: str) -> str:
    data = INPUT_KEYWORDS if role == "input" else OUTPUT_KEYWORDS
    return " | ".join(data.get(param, []))


def infer_unit(type_guess: str) -> str:
    if type_guess == "percent":
        return "decimal"
    if type_guess == "currency":
        return "VND"
    return ""


def scan_candidates(excel_path: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], dict[str, Any]]:
    wb = load_workbook(excel_path, data_only=False, read_only=False)
    input_candidates: list[dict[str, Any]] = []
    output_candidates: list[dict[str, Any]] = []
    error_rows: list[dict[str, Any]] = []
    structure = {
        "workbook": str(excel_path),
        "sheets": [{"title": ws.title, "max_row": ws.max_row, "max_column": ws.max_column} for ws in wb.worksheets],
    }

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if value is None:
                    continue
                label = nearby_label(ws, cell.row, cell.column)
                formula = is_formula(value)
                if contains_error(value):
                    error_rows.append({
                        "sheet": ws.title,
                        "cell": cell.coordinate,
                        "value_or_formula": compact_value(value),
                        "nearby_label": label,
                    })

                # Candidate input: hardcoded numeric/date cell, label nearby.
                if not formula and isinstance(value, (int, float, datetime)) and label:
                    param_guess = ""
                    for param, phrases in INPUT_KEYWORDS.items():
                        if text_matches(label, phrases):
                            param_guess = param
                            break
                    # Keep likely candidates. Cells with no param guess go to candidate sheet only, not review mapping.
                    confidence = guess_confidence(param_guess, label, "input", formula, value)
                    if param_guess or cell.column <= 15:
                        input_candidates.append({
                            "parameter_guess": param_guess,
                            "sheet": ws.title,
                            "cell": cell.coordinate,
                            "value": compact_value(value),
                            "type_guess": value_type_guess(value, label),
                            "is_formula": formula,
                            "nearby_label": label,
                            "confidence": confidence,
                        })

                # Candidate output: numeric or formula cell near output labels.
                if isinstance(value, (int, float)) or formula:
                    for key, phrases in OUTPUT_KEYWORDS.items():
                        if text_matches(label, phrases):
                            output_candidates.append({
                                "output_guess": key,
                                "sheet": ws.title,
                                "cell": cell.coordinate,
                                "value_or_formula": compact_value(value),
                                "type_guess": "percent" if "irr" in key or key == "roi" else "currency",
                                "is_formula": formula,
                                "nearby_label": label,
                                "confidence": guess_confidence(key, label, "output", formula, value),
                            })
                            break
    return input_candidates, output_candidates, error_rows, structure


def dedupe_review_rows(input_candidates: list[dict[str, Any]], output_candidates: list[dict[str, Any]]) -> list[list[Any]]:
    review_rows: list[list[Any]] = []
    seen_inputs: set[str] = set()
    for row in input_candidates:
        param = row.get("parameter_guess") or ""
        if not param or param in seen_inputs:
            continue
        seen_inputs.add(param)
        min_v, max_v = default_min_max(param, row.get("type_guess", ""))
        review_rows.append([
            "Review" if row.get("confidence") != "high" else "Approve",
            "input",
            param,
            FRIENDLY_NAMES.get(param, param.replace("_", " ").title()),
            row["sheet"],
            row["cell"],
            row.get("value", ""),
            "Yes" if row.get("is_formula") else "No",
            row.get("type_guess", "number"),
            infer_unit(row.get("type_guess", "")),
            min_v,
            max_v,
            "Yes",
            row.get("confidence", "medium"),
            row.get("nearby_label", ""),
            f"Candidate input detected near label: {row.get('nearby_label', '')}",
            aliases_for(param, "input"),
            "",
            "",
            "cell",
            "",
        ])

    seen_outputs: set[str] = set()
    for row in output_candidates:
        param = row.get("output_guess") or ""
        if not param or param in seen_outputs:
            continue
        seen_outputs.add(param)
        review_rows.append([
            "Review" if row.get("confidence") != "high" else "Approve",
            "output",
            param,
            FRIENDLY_NAMES.get(param, param.replace("_", " ").title()),
            row["sheet"],
            row["cell"],
            row.get("value_or_formula", ""),
            "Yes" if row.get("is_formula") else "No",
            row.get("type_guess", "number"),
            infer_unit(row.get("type_guess", "")),
            "",
            "",
            "No",
            row.get("confidence", "medium"),
            row.get("nearby_label", ""),
            f"Candidate output detected near label: {row.get('nearby_label', '')}",
            aliases_for(param, "output"),
            "",
            "",
            "cell",
            "",
        ])
    return review_rows


def write_table(ws, headers: list[str], rows: list[list[Any]], start_row: int = 1) -> None:
    ws.append(headers)
    for row in rows:
        ws.append(row)
    for cell in ws[start_row]:
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.font = Font(color=HEADER_FONT, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col_idx, header in enumerate(headers, start=1):
        width = 16
        if header in {"Nearby Label", "Description", "Aliases", "Notes"}:
            width = 36
        elif header in {"Parameter Key", "Friendly Name"}:
            width = 24
        elif header in {"Sheet"}:
            width = 22
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER


def style_review(ws) -> None:
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        role = row[1].value
        fill = INPUT_FILL if role == "input" else OUTPUT_FILL if role == "output" else None
        if fill:
            for cell in row:
                cell.fill = PatternFill("solid", fgColor=fill)
        # Highlight formula inputs if any.
        if role == "input" and row[7].value == "Yes":
            for cell in row:
                cell.fill = PatternFill("solid", fgColor=ERROR_FILL)
        # Highlight low confidence.
        if row[13].value == "low":
            row[13].fill = PatternFill("solid", fgColor=WARN_FILL)

    dv_decision = DataValidation(type="list", formula1='"Approve,Reject,Review"', allow_blank=False)
    dv_role = DataValidation(type="list", formula1='"input,output"', allow_blank=False)
    dv_yes_no = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
    dv_target = DataValidation(type="list", formula1='"cell,cells,range"', allow_blank=False)
    ws.add_data_validation(dv_decision)
    ws.add_data_validation(dv_role)
    ws.add_data_validation(dv_yes_no)
    ws.add_data_validation(dv_target)
    if ws.max_row >= 2:
        dv_decision.add(f"A2:A{ws.max_row}")
        dv_role.add(f"B2:B{ws.max_row}")
        dv_yes_no.add(f"M2:M{ws.max_row}")
        dv_target.add(f"T2:T{ws.max_row}")

    ws["A1"].comment = Comment("Approve rows that should become official mapping. Reject rows that are wrong. Use Correct Sheet / Correct Cell/Range to override system guesses.", "Excel AI Controller")
    ws["R1"].comment = Comment("Optional override. If blank, the Sheet column is used.", "Excel AI Controller")
    ws["S1"].comment = Comment("Optional override. Enter A1 cell, comma-separated cells, or a range. Target Mode decides how it is written to YAML.", "Excel AI Controller")


def create_review_workbook(excel_path: Path, out_path: Path) -> None:
    input_candidates, output_candidates, error_rows, structure = scan_candidates(excel_path)
    review_rows = dedupe_review_rows(input_candidates, output_candidates)

    wb = Workbook()
    ws = wb.active
    ws.title = "Review Mapping"
    write_table(ws, REVIEW_HEADERS, review_rows)
    style_review(ws)

    instructions = wb.create_sheet("Instructions", 0)
    instructions["A1"] = "Mapping Review Wizard"
    instructions["A1"].font = Font(size=18, bold=True, color="1F4E79")
    instructions["A3"] = "Purpose"
    instructions["B3"] = "Review the auto-detected Excel input/output mapping without editing YAML or code."
    instructions["A5"] = "How to use"
    steps = [
        "1. Open the 'Review Mapping' sheet.",
        "2. For each row, check whether the parameter/output matches the Sheet and Cell.",
        "3. Set Decision = Approve for correct rows, Reject for wrong rows, or Review if unsure.",
        "4. If the guessed cell is wrong, fill Correct Sheet and Correct Cell/Range.",
        "5. Save this workbook.",
        "6. Run scripts/build_profile_from_review.py to create config/profiles/<profile>/model_map.yaml.",
        "7. Run scripts/validate_profile.py before using the profile.",
    ]
    for idx, text in enumerate(steps, start=6):
        instructions[f"A{idx}"] = text
    instructions["A15"] = "Important rule"
    instructions["B15"] = "Inputs must normally be non-formula cells. Outputs can be formula cells. Never approve an input row that points to a formula unless you intentionally allow formula overwrite later."
    instructions.column_dimensions["A"].width = 24
    instructions.column_dimensions["B"].width = 100
    for row in instructions.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws_inputs = wb.create_sheet("Candidate Inputs")
    input_rows = [[r.get("parameter_guess", ""), r.get("sheet", ""), r.get("cell", ""), r.get("value", ""), r.get("type_guess", ""), "Yes" if r.get("is_formula") else "No", r.get("confidence", ""), r.get("nearby_label", "")] for r in input_candidates]
    write_table(ws_inputs, INPUT_HEADERS, input_rows)

    ws_outputs = wb.create_sheet("Candidate Outputs")
    output_rows = [[r.get("output_guess", ""), r.get("sheet", ""), r.get("cell", ""), r.get("value_or_formula", ""), r.get("type_guess", ""), "Yes" if r.get("is_formula") else "No", r.get("confidence", ""), r.get("nearby_label", "")] for r in output_candidates]
    write_table(ws_outputs, OUTPUT_HEADERS, output_rows)

    ws_errors = wb.create_sheet("Formula Errors")
    errors_out = [[r.get("sheet", ""), r.get("cell", ""), r.get("value_or_formula", ""), r.get("nearby_label", "")] for r in error_rows]
    write_table(ws_errors, ERROR_HEADERS, errors_out)
    if error_rows:
        for row in ws_errors.iter_rows(min_row=2):
            for cell in row:
                cell.fill = PatternFill("solid", fgColor=ERROR_FILL)

    ws_meta = wb.create_sheet("Workbook Structure")
    ws_meta["A1"] = "Workbook"
    ws_meta["B1"] = str(excel_path)
    ws_meta["A2"] = "Generated At"
    ws_meta["B2"] = datetime.now().isoformat(timespec="seconds")
    ws_meta["A4"] = "Sheet"
    ws_meta["B4"] = "Max Row"
    ws_meta["C4"] = "Max Column"
    for c in ws_meta[4]:
        c.fill = PatternFill("solid", fgColor=HEADER_FILL)
        c.font = Font(color=HEADER_FONT, bold=True)
    for item in structure["sheets"]:
        ws_meta.append([item["title"], item["max_row"], item["max_column"]])
    ws_meta.column_dimensions["A"].width = 36
    ws_meta.column_dimensions["B"].width = 16
    ws_meta.column_dimensions["C"].width = 16

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

    # Also write machine-readable summary beside the workbook.
    summary = {
        "excel": str(excel_path),
        "review_file": str(out_path),
        "review_rows": len(review_rows),
        "candidate_inputs": len(input_candidates),
        "candidate_outputs": len(output_candidates),
        "formula_errors": len(error_rows),
        "next_step": f"Edit {out_path.name}, approve rows, then run scripts/build_profile_from_review.py.",
    }
    (out_path.parent / "mapping_review_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate a human-readable Excel mapping review workbook.")
    parser.add_argument("--excel", required=True, help="Source Excel financial model path")
    parser.add_argument("--out", help="Output review .xlsx path. Defaults to outputs/analysis/<excel_stem>/mapping_review.xlsx")
    args = parser.parse_args()

    excel_path = Path(args.excel)
    if not excel_path.is_absolute():
        excel_path = ROOT / excel_path
    if not excel_path.exists():
        raise FileNotFoundError(excel_path)

    if args.out:
        out_path = Path(args.out)
        if not out_path.is_absolute():
            out_path = ROOT / out_path
    else:
        stem = safe_filename(excel_path.stem)
        out_path = ROOT / "outputs" / "analysis" / stem / "mapping_review.xlsx"

    create_review_workbook(excel_path, out_path)


if __name__ == "__main__":
    main()
