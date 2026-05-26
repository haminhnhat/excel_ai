from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries, get_column_letter

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from app.config_loader import load_model_map_for_profile


def input_targets(meta: dict[str, Any]) -> list[str]:
    if meta.get("cell"):
        return [str(meta["cell"])]
    if meta.get("cells"):
        return [str(c) for c in meta["cells"]]
    if meta.get("range"):
        min_col, min_row, max_col, max_row = range_boundaries(str(meta["range"]))
        return [f"{get_column_letter(c)}{r}" for r in range(min_row, max_row + 1) for c in range(min_col, max_col + 1)]
    return []


def is_formula(v: Any) -> bool:
    return isinstance(v, str) and v.startswith("=")


def main() -> None:
    parser = argparse.ArgumentParser(description="Validate a model profile against an Excel workbook.")
    parser.add_argument("--excel", required=True, help="Workbook path")
    parser.add_argument("--profile", help="Profile name under config/profiles/<name>")
    parser.add_argument("--map", help="Explicit model_map.yaml path")
    args = parser.parse_args()

    excel_path = Path(args.excel)
    if not excel_path.is_absolute():
        excel_path = ROOT / excel_path
    if not excel_path.exists():
        raise FileNotFoundError(excel_path)

    model_map, map_path = load_model_map_for_profile(profile=args.profile, explicit_path=args.map)
    wb = load_workbook(excel_path, data_only=False)

    errors: list[str] = []
    warnings: list[str] = []

    for key, meta in model_map.get("inputs", {}).items():
        sheet = meta.get("sheet")
        if sheet not in wb.sheetnames:
            errors.append(f"input {key}: sheet not found: {sheet}")
            continue
        targets = input_targets(meta)
        if not targets:
            errors.append(f"input {key}: must define cell, cells, or range")
            continue
        ws = wb[sheet]
        for cell in targets:
            try:
                v = ws[cell].value
            except Exception as e:
                errors.append(f"input {key}: invalid cell {sheet}!{cell}: {e}")
                continue
            if is_formula(v) and not meta.get("allow_formula_overwrite", False):
                errors.append(f"input {key}: mapped to formula cell {sheet}!{cell}: {v}")
            if v is None:
                warnings.append(f"input {key}: mapped cell is blank {sheet}!{cell}")

    for key, meta in model_map.get("outputs", {}).items():
        sheet = meta.get("sheet")
        cell = meta.get("cell")
        if sheet not in wb.sheetnames:
            errors.append(f"output {key}: sheet not found: {sheet}")
            continue
        try:
            _ = wb[sheet][cell].value
        except Exception as e:
            errors.append(f"output {key}: invalid cell {sheet}!{cell}: {e}")

    report = {
        "excel": str(excel_path),
        "model_map": str(map_path),
        "errors": errors,
        "warnings": warnings,
        "ok": not errors,
    }
    print(json.dumps(report, ensure_ascii=False, indent=2))
    if errors:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
