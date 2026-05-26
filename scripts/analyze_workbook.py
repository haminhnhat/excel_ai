from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

import yaml
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from app.utils import normalize_text, safe_filename

ERROR_TOKENS = ("#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#N/A", "#NUM!", "#NULL!")

OUTPUT_KEYWORDS = {
    "project_npv": ["npv du an", "npv dự án", "project npv", "npv"],
    "project_irr": ["irr du an", "irr dự án", "project irr", "irr"],
    "equity_npv": ["npv chu dau tu", "npv chủ đầu tư", "equity npv", "investor npv"],
    "equity_irr": ["irr chu dau tu", "irr chủ đầu tư", "equity irr", "investor irr"],
    "roi": ["roi", "return on investment"],
    "profit_after_tax": ["loi nhuan sau thue", "lợi nhuận sau thuế", "lnst", "profit after tax", "pat"],
    "total_investment": ["tong muc dau tu", "tổng mức đầu tư", "tmdt", "tmđt", "total investment"],
    "revenue": ["doanh thu", "revenue", "sales"],
    "bank_loan": ["vay ngan hang", "vay ngân hàng", "bank loan"],
}

INPUT_KEYWORDS = {
    "loan_interest_rate": ["lai vay", "lãi vay", "lai suat", "lãi suất", "interest rate", "chi phi von vay"],
    "vat_rate": ["vat", "thue vat", "thuế vat"],
    "cit_rate": ["tndn", "thuế tndn", "thue tndn", "cit"],
    "selling_price_change": ["gia ban", "giá bán", "don gia", "đơn giá", "selling price"],
    "investment_cost_change": ["tong muc dau tu", "tổng mức đầu tư", "tmdt", "tmđt", "chi phi dau tu"],
    "selling_cost_rate": ["chi phi ban hang", "chi phí bán hàng", "hoa hong", "moi gioi"],
    "marketing_rate": ["marketing", "truyen thong", "truyền thông"],
    "admin_cost_rate": ["chi phi quan ly", "chi phí quản lý", "g&a"],
}


def cell_addr(row: int, col: int) -> str:
    return f"{get_column_letter(col)}{row}"


def is_formula(value: Any) -> bool:
    return isinstance(value, str) and value.startswith("=")


def contains_error(value: Any) -> bool:
    if value is None:
        return False
    s = str(value)
    return any(tok in s for tok in ERROR_TOKENS)


def compact_value(value: Any) -> str:
    if value is None:
        return ""
    s = str(value)
    return s.replace("\n", " ").strip()[:300]


def text_matches(text: str, phrases: list[str]) -> bool:
    n = normalize_text(text or "")
    return any(normalize_text(p) in n for p in phrases)


def nearby_label(ws, row: int, col: int) -> str:
    labels: list[str] = []
    # Search left cells in the same row first. Most financial models use label-left/value-right.
    for c in range(max(1, col - 4), col):
        v = ws.cell(row=row, column=c).value
        if isinstance(v, str) and v.strip():
            labels.append(v.strip())
    # Then search above same column.
    for r in range(max(1, row - 3), row):
        v = ws.cell(row=r, column=col).value
        if isinstance(v, str) and v.strip():
            labels.append(v.strip())
    return " | ".join(labels)[:500]


def value_type_guess(value: Any, label: str = "") -> str:
    label_n = normalize_text(label)
    if isinstance(value, (int, float)):
        if -1 <= value <= 1 and any(k in label_n for k in ["ty le", "rate", "irr", "roi", "lai suat", "vat", "tndn", "percent"]):
            return "percent"
        if abs(value) > 1000:
            return "currency"
        return "number"
    if isinstance(value, datetime):
        return "date"
    return "text"


def sheet_bounds(ws) -> dict[str, Any]:
    return {
        "title": ws.title,
        "max_row": ws.max_row,
        "max_column": ws.max_column,
        "merged_ranges": [str(rng) for rng in ws.merged_cells.ranges][:100],
    }


def scan_workbook(excel_path: Path, out_dir: Path) -> None:
    out_dir.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(excel_path, data_only=False, read_only=False)

    structure = {
        "workbook": str(excel_path),
        "sheets": [sheet_bounds(ws) for ws in wb.worksheets],
        "defined_names": sorted([dn.name for dn in wb.defined_names.values()])[:500],
    }
    (out_dir / "workbook_structure.json").write_text(json.dumps(structure, ensure_ascii=False, indent=2), encoding="utf-8")

    formula_rows: list[dict[str, Any]] = []
    error_rows: list[dict[str, Any]] = []
    input_candidates: list[dict[str, Any]] = []
    output_candidates: list[dict[str, Any]] = []

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if value is None:
                    continue
                addr = cell.coordinate
                label = nearby_label(ws, cell.row, cell.column)

                if is_formula(value):
                    formula_rows.append({
                        "sheet": ws.title,
                        "cell": addr,
                        "formula": compact_value(value),
                        "nearby_label": label,
                    })
                if contains_error(value):
                    error_rows.append({
                        "sheet": ws.title,
                        "cell": addr,
                        "value_or_formula": compact_value(value),
                        "nearby_label": label,
                    })

                # Candidate inputs: hardcoded numeric/date cells with a nearby text label.
                if not is_formula(value) and isinstance(value, (int, float, datetime)) and label:
                    param_guess = ""
                    for param, phrases in INPUT_KEYWORDS.items():
                        if text_matches(label, phrases):
                            param_guess = param
                            break
                    if param_guess or cell.column <= 15:
                        input_candidates.append({
                            "sheet": ws.title,
                            "cell": addr,
                            "value": compact_value(value),
                            "type_guess": value_type_guess(value, label),
                            "nearby_label": label,
                            "parameter_guess": param_guess,
                            "is_formula": False,
                        })

                # Candidate outputs: if this numeric/formula cell is near an output label.
                if isinstance(value, (int, float)) or is_formula(value):
                    for key, phrases in OUTPUT_KEYWORDS.items():
                        if text_matches(label, phrases):
                            output_candidates.append({
                                "output_guess": key,
                                "sheet": ws.title,
                                "cell": addr,
                                "value_or_formula": compact_value(value),
                                "type_guess": "percent" if "irr" in key or key == "roi" else "currency" if key not in ["revenue"] else "currency",
                                "nearby_label": label,
                                "is_formula": is_formula(value),
                            })
                            break

    def write_csv(name: str, rows: list[dict[str, Any]]) -> None:
        path = out_dir / name
        if not rows:
            path.write_text("", encoding="utf-8-sig")
            return
        with path.open("w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
            writer.writeheader()
            writer.writerows(rows)

    write_csv("formula_cells.csv", formula_rows)
    write_csv("formula_errors.csv", error_rows)
    write_csv("candidate_inputs.csv", input_candidates)
    write_csv("candidate_outputs.csv", output_candidates)

    draft = build_draft_model_map(excel_path, input_candidates, output_candidates)
    (out_dir / "draft_model_map.yaml").write_text(yaml.safe_dump(draft, allow_unicode=True, sort_keys=False), encoding="utf-8")

    summary = {
        "excel": str(excel_path),
        "output_dir": str(out_dir),
        "sheet_count": len(wb.worksheets),
        "formula_cells": len(formula_rows),
        "formula_errors": len(error_rows),
        "candidate_inputs": len(input_candidates),
        "candidate_outputs": len(output_candidates),
        "next_step": "Open candidate_inputs.csv, candidate_outputs.csv and edit draft_model_map.yaml before using the profile.",
    }
    (out_dir / "analysis_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


def build_draft_model_map(excel_path: Path, input_candidates: list[dict[str, Any]], output_candidates: list[dict[str, Any]]) -> dict[str, Any]:
    # Keep only first candidate per guessed parameter/output. This is a draft, not trusted production mapping.
    inputs: dict[str, Any] = {}
    for row in input_candidates:
        key = row.get("parameter_guess")
        if not key or key in inputs:
            continue
        inputs[key] = {
            "sheet": row["sheet"],
            "cell": row["cell"],
            "type": row.get("type_guess") or "number",
            "editable": True,
            "min": -1.0 if "change" in key else 0.0,
            "max": 1.0,
            "unit": "decimal" if row.get("type_guess") == "percent" else None,
            "description": f"DRAFT candidate from nearby label: {row.get('nearby_label', '')}",
            "aliases": INPUT_KEYWORDS.get(key, []),
        }
    outputs: dict[str, Any] = {}
    for row in output_candidates:
        key = row.get("output_guess")
        if not key or key in outputs:
            continue
        outputs[key] = {
            "sheet": row["sheet"],
            "cell": row["cell"],
            "type": row.get("type_guess") or "number",
            "unit": "decimal" if row.get("type_guess") == "percent" else "VND" if row.get("type_guess") == "currency" else None,
            "description": f"DRAFT candidate from nearby label: {row.get('nearby_label', '')}",
            "aliases": OUTPUT_KEYWORDS.get(key, []),
        }

    return {
        "metadata": {
            "name": excel_path.stem,
            "version": "draft-0.1",
            "currency": "VND",
            "note": "DRAFT generated by scripts/analyze_workbook.py. Review every cell before use. Do not trust this mapping without human validation.",
        },
        "settings": {
            "allow_formula_overwrite": False,
            "default_outputs": list(outputs.keys())[:8],
        },
        "inputs": inputs,
        "outputs": outputs,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Analyze a financial Excel workbook and generate candidate mapping files.")
    parser.add_argument("--excel", required=True, help="Path to workbook")
    parser.add_argument("--out", default=None, help="Output folder. Default: outputs/analysis/<excel_stem>")
    args = parser.parse_args()

    excel_path = Path(args.excel)
    if not excel_path.is_absolute():
        excel_path = ROOT / excel_path
    if not excel_path.exists():
        raise FileNotFoundError(excel_path)

    out_dir = Path(args.out) if args.out else ROOT / "outputs" / "analysis" / safe_filename(excel_path.stem)
    if not out_dir.is_absolute():
        out_dir = ROOT / out_dir
    scan_workbook(excel_path, out_dir)


if __name__ == "__main__":
    main()
