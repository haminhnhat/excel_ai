from __future__ import annotations

import argparse
from pathlib import Path
from openpyxl import load_workbook


def main() -> None:
    parser = argparse.ArgumentParser(description="Find labels and formulas in an Excel workbook.")
    parser.add_argument("excel", help="Workbook path")
    parser.add_argument("--find", default="", help="Search text, e.g. NPV or Lợi nhuận")
    args = parser.parse_args()

    path = Path(args.excel)
    wb = load_workbook(path, data_only=False)
    print("Sheets:")
    for ws in wb.worksheets:
        print(f"- {ws.title}: {ws.max_row} rows x {ws.max_column} cols")

    if args.find:
        needle = args.find.lower()
        print(f"\nMatches for: {args.find}")
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    v = cell.value
                    if isinstance(v, str) and needle in v.lower():
                        print(f"{ws.title}!{cell.coordinate}: {v}")
    wb.close()


if __name__ == "__main__":
    main()
